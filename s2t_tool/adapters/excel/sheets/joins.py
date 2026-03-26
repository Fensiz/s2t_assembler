from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

from s2t_tool.shared.constants import JOIN_SOURCE_SQL
from s2t_tool.shared.files import read_json_file, read_text_file


def build_joins_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    sheet_name: str,
    joins_dir: str,
    diff_repo_dir: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    maybe_build_rich_diff: Callable[..., Any],
    join_row_key: Callable[[str, str], tuple[str, str]],
) -> None:
    diff_enabled = diff_repo_dir is not None
    root = repo_dir / joins_dir
    if not root.exists():
        raise ValueError(f"Required directory not found: {root}")

    old_rows_by_key: dict[tuple[str, str], dict[str, str]] = {}
    if diff_enabled:
        old_root = Path(diff_repo_dir) / joins_dir  # type: ignore[arg-type]
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                for old_load_dir in sorted([p for p in old_table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
                    join_json = read_json_file(old_load_dir / "join.json", default={}) or {}
                    key = join_row_key(old_table_dir.name, old_load_dir.name)
                    old_rows_by_key[key] = {
                        "description": join_json.get("description", ""),
                        "table_codes": "\n".join(join_json.get("table_codes", [])),
                        "table_codes_to_track_delta": "\n".join(join_json.get("table_codes_to_track_delta", [])),
                        "source_tables_join_sql": read_text_file(old_load_dir / JOIN_SOURCE_SQL),
                        "load_code_params": "\n".join(join_json.get("load_code_params", [])),
                        "settings_table_join_sql": read_text_file(old_load_dir / "settings_table_join.sql"),
                        "history_rule": join_json.get("history_rule", ""),
                        "business_history_dates": join_json.get("business_history_dates", ""),
                    }

    sheet = create_sheet(wb, sheet_name)
    sheet["B1"] = "TARGET"
    sheet["F1"] = "SOURCE"
    headers = [
        "Load code",
        "Table name",
        "Description",
        "Table code(s)",
        "Table codes to track delta",
        "Source tables join",
        "Load code params",
        "Settings table join",
        "History rule",
        "Business history dates",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    row_idx = 3
    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        for load_dir in sorted([p for p in table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
            join_json = read_json_file(load_dir / "join.json", default={}) or {}
            key = join_row_key(table_dir.name, load_dir.name)
            old_row = old_rows_by_key.get(key, {})

            current_description = join_json.get("description", "")
            current_table_codes = "\n".join(join_json.get("table_codes", []))
            current_delta_codes = "\n".join(join_json.get("table_codes_to_track_delta", []))
            current_source_join_sql = read_text_file(load_dir / JOIN_SOURCE_SQL)
            current_load_code_params = "\n".join(join_json.get("load_code_params", []))
            current_settings_join_sql = read_text_file(load_dir / "settings_table_join.sql")
            current_history_rule = join_json.get("history_rule", "")
            current_business_history_dates = join_json.get("business_history_dates", "")

            sheet.cell(row=row_idx, column=1, value=load_dir.name)
            sheet.cell(row=row_idx, column=2, value=table_dir.name)
            sheet.cell(row=row_idx, column=3, value=maybe_build_rich_diff(diff_enabled, old_row.get("description"), current_description))
            sheet.cell(row=row_idx, column=4, value=maybe_build_rich_diff(diff_enabled, old_row.get("table_codes"), current_table_codes))
            sheet.cell(row=row_idx, column=5, value=maybe_build_rich_diff(diff_enabled, old_row.get("table_codes_to_track_delta"), current_delta_codes))
            sheet.cell(row=row_idx, column=6, value=maybe_build_rich_diff(diff_enabled, old_row.get("source_tables_join_sql"), current_source_join_sql))
            sheet.cell(row=row_idx, column=7, value=maybe_build_rich_diff(diff_enabled, old_row.get("load_code_params"), current_load_code_params))
            sheet.cell(row=row_idx, column=8, value=maybe_build_rich_diff(diff_enabled, old_row.get("settings_table_join_sql"), current_settings_join_sql))
            sheet.cell(row=row_idx, column=9, value=maybe_build_rich_diff(diff_enabled, old_row.get("history_rule"), current_history_rule))
            sheet.cell(row=row_idx, column=10, value=maybe_build_rich_diff(diff_enabled, old_row.get("business_history_dates"), current_business_history_dates))
            row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)
