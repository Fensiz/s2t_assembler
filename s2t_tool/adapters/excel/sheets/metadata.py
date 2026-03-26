from __future__ import annotations

from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from s2t_tool.shared.resources import load_json_resource


def normalize_table_row(values: Any, headers: list[str] | None) -> list[Any]:
    if isinstance(values, list):
        return values
    if isinstance(values, dict):
        if headers:
            return [values.get(header, "") for header in headers]
        return list(values.values())
    return [values]


def build_metadata_sheet(
    wb: Workbook,
    config: dict[str, Any],
    sheet_name: str,
    metadata_json: str,
    diff_commit: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    apply_table_borders: Callable[..., None],
) -> None:
    payload = load_json_resource(metadata_json) or {}
    blocks = payload.get("blocks", [])
    sheet = create_sheet(wb, sheet_name)
    row_idx = 1

    if diff_commit:
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        cell = sheet.cell(row=row_idx, column=1, value=f"DIFF FILE — DO NOT USE FOR PUT (compared with commit: {diff_commit})")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_idx += 2

    for block in blocks:
        title_en = block.get("title_en")
        title_ru = block.get("title_ru")
        link = block.get("link")
        table = block.get("table")

        if title_en:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=title_en)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

        if title_ru:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=title_ru)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

        if link:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=link)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        for text_line in block.get("lines", []) or []:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=text_line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        if table:
            headers = table.get("headers")
            rows = table.get("rows", []) or []
            table_start_row = row_idx
            header_names: list[str] | None = None

            if headers:
                header_names = [str(h) for h in headers]
                for col_idx, header in enumerate(header_names, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row_idx += 1

            max_width = len(header_names) if header_names else 0
            for raw_row in rows:
                normalized_row = normalize_table_row(raw_row, header_names)
                max_width = max(max_width, len(normalized_row))
                for col_idx, value in enumerate(normalized_row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1

            table_end_row = row_idx - 1
            if max_width > 0 and table_end_row >= table_start_row:
                apply_table_borders(sheet, table_start_row, table_end_row, 1, max_width)

        row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)
