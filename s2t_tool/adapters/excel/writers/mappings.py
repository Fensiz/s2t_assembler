from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

from s2t_tool.shared.csv_files import read_csv_rows
from s2t_tool.shared.files import read_json_file
from s2t_tool.shared.text import normalize_newlines


def resolve_attribute_name(table_name: str, attribute_code: str, attribute_names: dict[str, Any]) -> str:
    tables = attribute_names.get("tables", {}) or {}
    common = attribute_names.get("common", {}) or {}
    if table_name in tables and attribute_code in tables[table_name]:
        return tables[table_name][attribute_code]
    return common.get(attribute_code, "")


def build_mappings_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    sheet_name: str,
    joins_dir: str,
    attribute_names_json: str,
    mappings_csv: str,
    mappings_extra_json: str,
    diff_repo_dir: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    maybe_build_rich_diff: Callable[..., Any],
    mapping_row_key: Callable[[str, str, str], tuple[str, str, str]],
) -> None:
    diff_enabled = diff_repo_dir is not None
    old_rows_by_key: dict[tuple[str, str, str], dict[str, str]] = {}

    if diff_enabled:
        old_root = Path(diff_repo_dir) / joins_dir  # type: ignore[arg-type]
        old_attribute_names = read_json_file(Path(diff_repo_dir) / attribute_names_json, default={}) or {}
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                for old_load_dir in sorted([p for p in old_table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
                    mappings_path = old_load_dir / mappings_csv
                    if not mappings_path.exists():
                        continue

                    _, old_rows = read_csv_rows(mappings_path)
                    old_extra = read_json_file(old_load_dir / mappings_extra_json, default={}) or {}
                    for row in old_rows:
                        while len(row) < 2:
                            row.append("")
                        attribute_code = row[0]
                        extra_item = old_extra.get(attribute_code, {}) or {}
                        key = mapping_row_key(old_load_dir.name, old_table_dir.name, attribute_code)
                        old_rows_by_key[key] = {
                            "attribute_name": resolve_attribute_name(old_table_dir.name, attribute_code, old_attribute_names),
                            "mapping_algorithm": row[1],
                            "additional_join": extra_item.get("additional_join", ""),
                            "settings": extra_item.get("settings", ""),
                        }

    root = repo_dir / joins_dir
    attribute_names = read_json_file(repo_dir / attribute_names_json, default={}) or {}
    sheet = create_sheet(wb, sheet_name)
    sheet.cell(row=1, column=2, value="TARGET")
    sheet.cell(row=1, column=5, value="SOURCE")

    headers = [
        "Load code",
        "Table name",
        "Attribute code",
        "Attribute name",
        "Mapping algorithm",
        "Additional join",
        "Settings",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    row_idx = 3
    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        for load_dir in sorted([p for p in table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
            mappings_path = load_dir / mappings_csv
            if not mappings_path.exists():
                continue

            _, rows = read_csv_rows(mappings_path)
            extra = read_json_file(load_dir / mappings_extra_json, default={}) or {}
            for row in rows:
                while len(row) < 2:
                    row.append("")
                attribute_code = row[0]
                mapping_algorithm = normalize_newlines(row[1])
                extra_item = extra.get(attribute_code, {}) or {}

                sheet.cell(row=row_idx, column=1, value=load_dir.name)
                sheet.cell(row=row_idx, column=2, value=table_dir.name)
                sheet.cell(row=row_idx, column=3, value=attribute_code)

                key = mapping_row_key(load_dir.name, table_dir.name, attribute_code)
                old_row = old_rows_by_key.get(key, {})
                current_attribute_name = resolve_attribute_name(table_dir.name, attribute_code, attribute_names)
                current_additional_join = extra_item.get("additional_join", "")
                current_settings = extra_item.get("settings", "")

                sheet.cell(row=row_idx, column=4, value=maybe_build_rich_diff(diff_enabled, old_row.get("attribute_name"), current_attribute_name))
                sheet.cell(row=row_idx, column=5, value=maybe_build_rich_diff(diff_enabled, old_row.get("mapping_algorithm"), mapping_algorithm))
                sheet.cell(row=row_idx, column=6, value=maybe_build_rich_diff(diff_enabled, old_row.get("additional_join"), current_additional_join))
                sheet.cell(row=row_idx, column=7, value=maybe_build_rich_diff(diff_enabled, old_row.get("settings"), current_settings))
                row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)
