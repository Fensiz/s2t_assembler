from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

from s2t_tool.shared.files import read_json_file, read_text_file


def build_pre_transforms_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    sheet_name: str,
    pre_transforms_dir: str,
    diff_repo_dir: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    maybe_build_rich_diff: Callable[..., Any],
    pre_transform_row_key: Callable[[str], str],
) -> None:
    diff_enabled = diff_repo_dir is not None

    sheet = create_sheet(wb, sheet_name)
    sheet.cell(row=1, column=1, value="TARGET")
    sheet.cell(row=1, column=3, value="SOURCE")

    headers = [
        "Target table",
        "Source tables",
        "Transformation SQL",
        "Comment",
        "Settings SQL",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    root = repo_dir / pre_transforms_dir
    if not root.exists():
        finalize_sheet_style(sheet, config, sheet_name)
        return

    old_rows_by_key: dict[str, dict[str, str]] = {}
    if diff_enabled:
        old_root = Path(diff_repo_dir) / pre_transforms_dir  # type: ignore[arg-type]
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                json_path = old_table_dir / "pre-transform.json"
                if not json_path.exists():
                    continue

                payload = read_json_file(json_path, default={}) or {}
                target_table = payload.get("target_table", old_table_dir.name)
                old_rows_by_key[pre_transform_row_key(target_table)] = {
                    "source_tables": "\n".join(payload.get("source_tables", [])),
                    "comments": payload.get("comments", ""),
                    "transformation_sql": read_text_file(old_table_dir / "preliminary_transformation.sql"),
                    "settings_sql": read_text_file(old_table_dir / "settings.sql"),
                }

    row_idx = 3
    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        json_path = table_dir / "pre-transform.json"
        if not json_path.exists():
            continue

        payload = read_json_file(json_path, default={}) or {}
        target_table = payload.get("target_table", table_dir.name)
        source_tables = "\n".join(payload.get("source_tables", []))
        comments = payload.get("comments", "")
        transformation_sql = read_text_file(table_dir / "preliminary_transformation.sql")
        settings_sql = read_text_file(table_dir / "settings.sql")

        key = pre_transform_row_key(target_table)
        old_row = old_rows_by_key.get(key, {})

        sheet.cell(row=row_idx, column=1, value=target_table)
        sheet.cell(row=row_idx, column=2, value=maybe_build_rich_diff(diff_enabled, old_row.get("source_tables"), source_tables))
        sheet.cell(row=row_idx, column=3, value=maybe_build_rich_diff(diff_enabled, old_row.get("transformation_sql"), transformation_sql))
        sheet.cell(row=row_idx, column=4, value=maybe_build_rich_diff(diff_enabled, old_row.get("comments"), comments))
        sheet.cell(row=row_idx, column=5, value=maybe_build_rich_diff(diff_enabled, old_row.get("settings_sql"), settings_sql))
        row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)
