from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from s2t_tool.adapters.excel.reader_sections import (
    export_change_history as export_change_history_section,
    export_joins as export_joins_section,
    export_mappings as export_mappings_section,
    export_pre_transforms as export_pre_transforms_section,
    export_simple_csv_sheet as export_simple_csv_sheet_section,
    export_source_lg as export_source_lg_section,
    export_targets as export_targets_section,
)
from s2t_tool.domain.schema import DEFAULT_SCHEMA, S2TSchema


class ExcelRepoReader:
    def __init__(
        self,
        excel_path: str | Path,
        output_dir: str | Path,
        schema: S2TSchema = DEFAULT_SCHEMA,
        format_sql: bool = False,
        logger=None,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.schema = schema
        self.format_sql = format_sql
        self.logger = logger
        self.workbook: Workbook = load_workbook(self.excel_path, data_only=True)

    def get_sheet(self, sheet_name: str, required: bool = True) -> Worksheet | None:
        normalized_aliases = {
            self.schema.normalize_sheet_name(alias)
            for alias in self.schema.sheet_aliases_for(sheet_name)
        }

        for actual_name in self.workbook.sheetnames:
            if self.schema.normalize_sheet_name(actual_name) in normalized_aliases:
                return self.workbook[actual_name]

        if required:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {self.excel_path}. "
                f"Available sheets: {self.workbook.sheetnames}"
            )

        return None

    def export_change_history(self) -> Path:
        sheet = self.get_sheet(self.schema.change_history_sheet, required=True)
        assert sheet is not None
        return export_change_history_section(self.output_dir, self.schema, sheet)

    def export_source_lg(self) -> Path:
        sheet = self.get_sheet(self.schema.source_lg_sheet, required=True)
        assert sheet is not None
        return export_source_lg_section(self.output_dir, self.schema, sheet)

    def export_targets(self) -> Path:
        sheet = self.get_sheet(self.schema.targets_sheet, required=True)
        assert sheet is not None
        return export_targets_section(self.output_dir, self.schema, sheet)

    def export_simple_csv_sheet(
        self,
        sheet_name: str,
        output_name: str,
        headers: list[str],
        required: bool = False,
    ) -> Path | None:
        sheet = self.get_sheet(sheet_name, required=required)
        return export_simple_csv_sheet_section(self.output_dir, output_name, headers, sheet)

    def export_pre_transforms(self) -> Path:
        sheet = self.get_sheet(self.schema.pre_transforms_sheet, required=True)
        assert sheet is not None
        return export_pre_transforms_section(self.output_dir, self.schema, sheet, self.format_sql)

    def export_joins(self) -> Path:
        sheet = self.get_sheet(self.schema.joins_sheet, required=True)
        assert sheet is not None
        return export_joins_section(self.output_dir, self.schema, sheet, self.format_sql)

    def export_mappings(self) -> tuple[Path, Path]:
        sheet = self.get_sheet(self.schema.mappings_sheet, required=True)
        assert sheet is not None
        return export_mappings_section(self.output_dir, self.schema, sheet)

    def export_all(self) -> None:
        created: list[Path] = [
            self.export_change_history(),
            self.export_source_lg(),
            self.export_targets(),
            self.export_pre_transforms(),
            self.export_joins(),
        ]

        mappings_root, attribute_names_path = self.export_mappings()
        created.append(mappings_root)
        created.append(attribute_names_path)

        for maybe_path in [
            self.export_simple_csv_sheet(
                self.schema.settings_sheet,
                self.schema.settings_csv,
                list(self.schema.settings_headers),
                required=False,
            ),
            self.export_simple_csv_sheet(
                self.schema.parameters_sheet,
                self.schema.parameters_csv,
                list(self.schema.parameters_headers),
                required=False,
            ),
            self.export_simple_csv_sheet(
                self.schema.st_decoder_sheet,
                self.schema.st_decoder_csv,
                list(self.schema.settings_headers),
                required=False,
            ),
            self.export_simple_csv_sheet(
                self.schema.st_filter_sheet,
                self.schema.st_filter_csv,
                list(self.schema.st_filter_headers),
                required=False,
            ),
        ]:
            if maybe_path is not None:
                created.append(maybe_path)

        for path in created:
            if self.logger:
                self.logger(f"Created: {path}")


def normalize_sheet_name(value: str) -> str:
    return DEFAULT_SCHEMA.normalize_sheet_name(value)


def get_sheet(excel_path: Path, sheet_name: str, required: bool = True) -> Worksheet | None:
    reader = ExcelRepoReader(excel_path=excel_path, output_dir=excel_path.parent)
    return reader.get_sheet(sheet_name, required=required)


def export_change_history(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_change_history()


def export_source_lg(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_source_lg()


def export_targets(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_targets()


def export_simple_csv_sheet(
    excel_path: Path,
    output_dir: Path,
    sheet_name: str,
    output_name: str,
    headers: list[str],
    required: bool = False,
) -> Path | None:
    return ExcelRepoReader(excel_path, output_dir).export_simple_csv_sheet(
        sheet_name=sheet_name,
        output_name=output_name,
        headers=headers,
        required=required,
    )


def export_pre_transforms(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_pre_transforms()


def export_joins(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_joins()


def export_mappings(excel_path: Path, output_dir: Path) -> tuple[Path, Path]:
    return ExcelRepoReader(excel_path, output_dir).export_mappings()


def section_key_from_title(title: str) -> str:
    return (
        title.strip()
        .lower()
        .replace("-", " ")
        .replace("/", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace(":", "")
        .replace(".", "")
        .replace(",", "")
        .replace("  ", " ")
        .replace(" ", "_")
    )


def export_excel_to_repo(excel_path: str, output_dir: str, format_sql: bool = False, logger=None) -> None:
    ExcelRepoReader(
        excel_path=excel_path,
        output_dir=output_dir,
        schema=DEFAULT_SCHEMA,
        format_sql=format_sql,
        logger=logger,
    ).export_all()


if __name__ == "__main__":
    export_excel_to_repo("input.xlsx", "./repo")
