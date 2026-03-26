from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from s2t_tool.adapters.excel.writers import (
    build_change_history_sheet as build_change_history_sheet_sheet,
    build_joins_sheet as build_joins_sheet_sheet,
    build_mappings_sheet as build_mappings_sheet_sheet,
    build_metadata_sheet as build_metadata_sheet_sheet,
    build_pre_transforms_sheet as build_pre_transforms_sheet_sheet,
    build_source_lg_sheet as build_source_lg_sheet_sheet,
    build_targets_sheet as build_targets_sheet_sheet,
)
from s2t_tool.domain.schema import DEFAULT_SCHEMA, S2TSchema
from s2t_tool.adapters.excel.writer_style import (
    WRITER_CONFIG_FILE,
    apply_table_borders,
    append_csv_sheet,
    create_sheet,
    finalize_sheet_style,
    load_writer_config,
)
from s2t_tool.adapters.excel.writer_diff import (
    build_rich_diff,
    maybe_build_rich_diff,
)
from s2t_tool.adapters.excel.writer_diff import join_row_key, mapping_row_key, pre_transform_row_key


SCHEMA = DEFAULT_SCHEMA

CHANGE_HISTORY_JSON = SCHEMA.change_history_json
SOURCE_LG_CSV = SCHEMA.source_lg_csv
TARGETS_CSV = SCHEMA.targets_csv
SETTINGS_CSV = SCHEMA.settings_csv
PARAMETERS_CSV = SCHEMA.parameters_csv
ST_DECODER_CSV = SCHEMA.st_decoder_csv
ST_FILTER_CSV = SCHEMA.st_filter_csv
METADATA_JSON = SCHEMA.metadata_json
ATTRIBUTE_NAMES_JSON = SCHEMA.attribute_names_json

PRE_TRANSFORMS_DIR = SCHEMA.pre_transforms_dir
JOINS_DIR = SCHEMA.joins_dir

CHANGE_HISTORY_SHEET = SCHEMA.change_history_sheet
SOURCE_LG_SHEET = SCHEMA.source_lg_sheet
TARGETS_SHEET = SCHEMA.targets_sheet
PRE_TRANSFORMS_SHEET = SCHEMA.pre_transforms_sheet
JOINS_SHEET = SCHEMA.joins_sheet
MAPPINGS_SHEET = SCHEMA.mappings_sheet
SETTINGS_SHEET = SCHEMA.settings_sheet
PARAMETERS_SHEET = SCHEMA.parameters_sheet
ST_DECODER_SHEET = SCHEMA.st_decoder_sheet
ST_FILTER_SHEET = SCHEMA.st_filter_sheet
METADATA_SHEET = SCHEMA.metadata_sheet


def build_change_history_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_change_history_sheet_sheet(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        change_history_json=CHANGE_HISTORY_JSON,
        change_history_sheet=CHANGE_HISTORY_SHEET,
        pre_transforms_sheet=PRE_TRANSFORMS_SHEET,
        joins_sheet=JOINS_SHEET,
        mappings_sheet=MAPPINGS_SHEET,
        diff_repo_dir=diff_repo_dir,
    )


def build_source_lg_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_source_lg_sheet_sheet(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        source_lg_sheet=SOURCE_LG_SHEET,
        source_lg_csv=SOURCE_LG_CSV,
        pre_transforms_sheet=PRE_TRANSFORMS_SHEET,
        joins_sheet=JOINS_SHEET,
        mappings_sheet=MAPPINGS_SHEET,
        diff_repo_dir=diff_repo_dir,
    )


def build_targets_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_targets_sheet_sheet(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        targets_sheet=TARGETS_SHEET,
        targets_csv=TARGETS_CSV,
        pre_transforms_sheet=PRE_TRANSFORMS_SHEET,
        joins_sheet=JOINS_SHEET,
        mappings_sheet=MAPPINGS_SHEET,
        diff_repo_dir=diff_repo_dir,
    )


# ============================================================
# Pre-transforms
# ============================================================

def build_pre_transforms_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_pre_transforms_sheet_sheet(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        sheet_name=PRE_TRANSFORMS_SHEET,
        pre_transforms_dir=PRE_TRANSFORMS_DIR,
        diff_repo_dir=diff_repo_dir,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        maybe_build_rich_diff=maybe_build_rich_diff,
        pre_transform_row_key=pre_transform_row_key,
    )


# ============================================================
# Joins
# ============================================================

def build_joins_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_joins_sheet_sheet(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        sheet_name=JOINS_SHEET,
        joins_dir=JOINS_DIR,
        diff_repo_dir=diff_repo_dir,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        maybe_build_rich_diff=maybe_build_rich_diff,
        join_row_key=join_row_key,
    )


# ============================================================
# Mappings
# ============================================================

def build_mappings_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_mappings_sheet_sheet(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        sheet_name=MAPPINGS_SHEET,
        joins_dir=JOINS_DIR,
        attribute_names_json=ATTRIBUTE_NAMES_JSON,
        mappings_csv="mappings.csv",
        mappings_extra_json="mappings.extra.json",
        diff_repo_dir=diff_repo_dir,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        maybe_build_rich_diff=maybe_build_rich_diff,
        mapping_row_key=mapping_row_key,
    )


# ============================================================
# Metadata
# ============================================================

def build_metadata_sheet(
    wb: Workbook,
    config: dict[str, Any],
    diff_commit: str | None = None,
) -> None:
    build_metadata_sheet_sheet(
        wb=wb,
        config=config,
        sheet_name=METADATA_SHEET,
        metadata_json=METADATA_JSON,
        diff_commit=diff_commit,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        apply_table_borders=apply_table_borders,
    )


# ============================================================
# Main entry
# ============================================================

def build_excel_from_repo(
    repo_dir: str,
    output_excel_path: str,
    config_path: str = WRITER_CONFIG_FILE,
    diff_repo_dir: str | None = None,
    diff_commit: str | None = None,
    logger=None,
) -> None:
    RepoExcelWriter(
        repo_dir=repo_dir,
        output_excel_path=output_excel_path,
        config_path=config_path,
        diff_repo_dir=diff_repo_dir,
        diff_commit=diff_commit,
        logger=logger,
        schema=DEFAULT_SCHEMA,
    ).build()


class RepoExcelWriter:
    def __init__(
        self,
        repo_dir: str | Path,
        output_excel_path: str | Path,
        config_path: str | Path = WRITER_CONFIG_FILE,
        diff_repo_dir: str | Path | None = None,
        diff_commit: str | None = None,
        logger=None,
        schema: S2TSchema = DEFAULT_SCHEMA,
    ) -> None:
        self.repo_dir = Path(repo_dir)
        self.output_excel_path = Path(output_excel_path)
        self.config = load_writer_config(config_path)
        self.diff_repo_dir = str(diff_repo_dir) if diff_repo_dir is not None else None
        self.diff_commit = diff_commit
        self.logger = logger
        self.schema = schema

    def build(self) -> None:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        self.build_change_history_sheet(wb)
        self.build_source_lg_sheet(wb)
        self.build_targets_sheet(wb)
        self.build_pre_transforms_sheet(wb)
        self.build_joins_sheet(wb)
        self.build_mappings_sheet(wb)
        self.append_optional_csv_sheets(wb)
        self.build_metadata_sheet(wb)

        wb.save(self.output_excel_path)
        if self.logger:
            self.logger(f"Created: {self.output_excel_path}")

    def build_change_history_sheet(self, wb: Workbook) -> None:
        build_change_history_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_source_lg_sheet(self, wb: Workbook) -> None:
        build_source_lg_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_targets_sheet(self, wb: Workbook) -> None:
        build_targets_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_pre_transforms_sheet(self, wb: Workbook) -> None:
        build_pre_transforms_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_joins_sheet(self, wb: Workbook) -> None:
        build_joins_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_mappings_sheet(self, wb: Workbook) -> None:
        build_mappings_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def append_optional_csv_sheets(self, wb: Workbook) -> None:
        append_csv_sheet(
            wb, self.schema.settings_sheet, self.repo_dir / self.schema.settings_csv, self.config, False,
            PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET,
            diff_csv_path=(Path(self.diff_repo_dir) / self.schema.settings_csv) if self.diff_repo_dir else None,
            maybe_build_rich_diff=maybe_build_rich_diff,
        )
        append_csv_sheet(
            wb, self.schema.parameters_sheet, self.repo_dir / self.schema.parameters_csv, self.config, False,
            PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET,
            diff_csv_path=(Path(self.diff_repo_dir) / self.schema.parameters_csv) if self.diff_repo_dir else None,
            maybe_build_rich_diff=maybe_build_rich_diff,
        )
        append_csv_sheet(
            wb, self.schema.st_decoder_sheet, self.repo_dir / self.schema.st_decoder_csv, self.config, False,
            PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET,
            diff_csv_path=(Path(self.diff_repo_dir) / self.schema.st_decoder_csv) if self.diff_repo_dir else None,
            maybe_build_rich_diff=maybe_build_rich_diff,
        )
        append_csv_sheet(
            wb, self.schema.st_filter_sheet, self.repo_dir / self.schema.st_filter_csv, self.config, False,
            PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET,
            diff_csv_path=(Path(self.diff_repo_dir) / self.schema.st_filter_csv) if self.diff_repo_dir else None,
            maybe_build_rich_diff=maybe_build_rich_diff,
        )

    def build_metadata_sheet(self, wb: Workbook) -> None:
        build_metadata_sheet(wb, self.config, self.diff_commit)


if __name__ == "__main__":
    build_excel_from_repo("./repo", "output.xlsx", WRITER_CONFIG_FILE)
