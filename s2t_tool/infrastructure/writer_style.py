from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from s2t_tool.shared.csv_files import read_csv_rows
from s2t_tool.shared.resources import load_json_resource


WRITER_CONFIG_FILE = "writer_config.json"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

EXCEL_HEADER_LABELS = {
    "scheme": "Scheme",
    "table": "Table",
    "column": "Column",
    "data_type": "Data type",
    "data_length": "Data length",
    "is_key": "Is a key",
    "description": "Description",
    "link": "Link",
    "table_code": "Table code",
    "table_name": "Table name",
    "settings_alias": "Settings alias",
    "settings_description": "Settings description",
    "settings_table": "Settings table",
    "settings_type": "Settings type",
    "period": "Period",
    "mask": "Mask",
    "parameter": "Parameter",
    "value": "Value",
    "comment": "Comment",
    "source_value": "Source value",
    "target_value": "Target value",
    "start_dt": "Start dt",
    "end_dt": "End dt",
    "update_date": "Update date",
    "author_update_name": "Author update name",
    "author": "Author",
    "date": "Date",
    "version": "Version",
    "jira_ticket": "Jira ticket",
    "load_code": "Load code",
    "attribute_code": "Attribute code",
    "attribute_name": "Attribute name",
    "mapping_algorithm": "Mapping algorithm",
    "additional_join": "Additional join",
    "business_history_dates": "Business history dates",
    "load_code_params": "Load code params",
    "table_codes_to_track_delta": "Table codes to track delta",
    "source_tables_join": "Source tables join",
    "settings_table_join": "Settings table join",
    "history_rule": "History rule",
}


def repo_header_to_excel(header: str) -> str:
    return EXCEL_HEADER_LABELS.get(header, header.replace("_", " ").capitalize())


def load_writer_config(config_path: str | Path | None = None) -> dict[str, Any]:
    if config_path is None:
        return load_json_resource(WRITER_CONFIG_FILE)

    path = Path(config_path)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))

    return load_json_resource(str(config_path))


def create_sheet(wb: Workbook, title: str) -> Worksheet:
    return wb.create_sheet(title=title)


def sheet_config(config: dict[str, Any], sheet_name: str) -> dict[str, Any]:
    global_cfg = config.get("global", {})
    per_sheet = config.get("sheets", {}).get(sheet_name, {})
    return {**global_cfg, **per_sheet}


def color_fill(hex_color: str | None) -> PatternFill | None:
    if not hex_color:
        return None
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)


def center_merged_headers(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    for merge_range in cfg.get("header_merge_ranges", []):
        first_cell = merge_range.split(":")[0]
        cell = sheet[first_cell]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_base_style(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    font_name = cfg.get("font_name", "Calibri")
    font_size = cfg.get("font_size", 11)
    wrap_text = cfg.get("wrap_text", True)
    vertical_alignment = cfg.get("vertical_alignment", "top")

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name=font_name, size=font_size)
            cell.alignment = Alignment(wrap_text=wrap_text, vertical=vertical_alignment)


def apply_header_style(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    header_rows = cfg.get("header_rows", 1)
    fill = color_fill(cfg.get("header_fill_color"))
    font_name = cfg.get("font_name", "Calibri")
    font_size = cfg.get("font_size", 11)

    for row_idx in range(1, header_rows + 1):
        for cell in sheet[row_idx]:
            cell.font = Font(name=font_name, size=font_size, bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill


def estimate_row_height_for_text(
    values: list[str],
    min_height: float,
    max_height: float,
    line_height: float = 15.0,
) -> float:
    max_lines = 1
    for value in values:
        if not value:
            continue
        max_lines = max(max_lines, str(value).count("\n") + 1)

    estimated = max(min_height, max_lines * line_height)
    return min(estimated, max_height)


def row_has_multiline_values(sheet: Worksheet, row_idx: int) -> bool:
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=row_idx, column=col_idx).value
        if value is not None and "\n" in str(value):
            return True
    return False


def row_values_as_strings(sheet: Worksheet, row_idx: int) -> list[str]:
    values: list[str] = []
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=row_idx, column=col_idx).value
        values.append("" if value is None else str(value))
    return values


def apply_row_heights(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    header_rows = cfg.get("header_rows", 1)
    header_height = cfg.get("header_row_height", 24)

    row_height_mode = cfg.get("row_height_mode", "fixed")
    body_row_height = cfg.get("body_row_height", 36)
    auto_height_on_multiline = cfg.get("auto_height_on_multiline", False)
    min_body_row_height = cfg.get("min_body_row_height", body_row_height)
    max_body_row_height = cfg.get("max_body_row_height", 180)

    for row_idx in range(1, sheet.max_row + 1):
        if row_idx <= header_rows:
            sheet.row_dimensions[row_idx].height = header_height
            continue

        if row_height_mode == "auto":
            values = row_values_as_strings(sheet, row_idx)
            sheet.row_dimensions[row_idx].height = estimate_row_height_for_text(
                values=values,
                min_height=min_body_row_height,
                max_height=max_body_row_height,
            )
            continue

        if auto_height_on_multiline and row_has_multiline_values(sheet, row_idx):
            values = row_values_as_strings(sheet, row_idx)
            sheet.row_dimensions[row_idx].height = estimate_row_height_for_text(
                values=values,
                min_height=min_body_row_height,
                max_height=max_body_row_height,
            )
        else:
            sheet.row_dimensions[row_idx].height = body_row_height


def apply_column_widths(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    widths = cfg.get("column_widths", {})
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width


def apply_merge_ranges(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    for merge_range in cfg.get("header_merge_ranges", []):
        sheet.merge_cells(merge_range)


def apply_freeze_panes(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    freeze_panes = cfg.get("freeze_panes")
    if freeze_panes:
        sheet.freeze_panes = freeze_panes


def apply_autofilter(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    if not cfg.get("add_autofilter", True):
        return

    if sheet.max_column == 0 or sheet.max_row == 0:
        return

    autofilter_row = cfg.get("autofilter_row", cfg.get("header_rows", 1))
    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A{autofilter_row}:{last_col}{sheet.max_row}"


def fill_row(sheet: Worksheet, row_idx: int, color_hex: str | None) -> None:
    fill = color_fill(color_hex)
    if fill is None:
        return

    for col_idx in range(1, sheet.max_column + 1):
        sheet.cell(row=row_idx, column=col_idx).fill = fill


def apply_special_header_colors(
    sheet: Worksheet,
    sheet_name: str,
    cfg: dict[str, Any],
    pre_transforms_sheet: str,
    joins_sheet: str,
    mappings_sheet: str,
) -> None:
    group_colors = cfg.get("group_colors", {})
    if not group_colors:
        return

    loadcode_fill = color_fill(group_colors.get("loadcode_header"))
    target_fill = color_fill(group_colors.get("target_section"))
    source_fill = color_fill(group_colors.get("source_section"))

    if sheet_name == pre_transforms_sheet:
        for col in range(1, 3):
            sheet.cell(row=1, column=col).fill = target_fill
        for col in range(3, 6):
            sheet.cell(row=1, column=col).fill = source_fill
        for col in range(1, 3):
            sheet.cell(row=2, column=col).fill = target_fill
        for col in range(3, 6):
            sheet.cell(row=2, column=col).fill = source_fill
    elif sheet_name == joins_sheet:
        sheet.cell(row=1, column=1).fill = loadcode_fill
        for col in range(2, 6):
            sheet.cell(row=1, column=col).fill = target_fill
        for col in range(6, 11):
            sheet.cell(row=1, column=col).fill = source_fill
        sheet.cell(row=2, column=1).fill = loadcode_fill
        for col in range(2, 6):
            sheet.cell(row=2, column=col).fill = target_fill
        for col in range(6, 11):
            sheet.cell(row=2, column=col).fill = source_fill
    elif sheet_name == mappings_sheet:
        sheet.cell(row=1, column=1).fill = loadcode_fill
        for col in range(2, 5):
            sheet.cell(row=1, column=col).fill = target_fill
        for col in range(5, 8):
            sheet.cell(row=1, column=col).fill = source_fill
        sheet.cell(row=2, column=1).fill = loadcode_fill
        for col in range(2, 5):
            sheet.cell(row=2, column=col).fill = target_fill
        for col in range(5, 8):
            sheet.cell(row=2, column=col).fill = source_fill


def apply_cell_borders(sheet: Worksheet) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def apply_table_borders(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: int = 2,
) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row=row_idx, column=col_idx).border = THIN_BORDER


def apply_alternating_group_fill(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    grouping = cfg.get("grouping", {})
    if not grouping.get("enabled", False):
        return
    group_by_column = grouping.get("group_by_column")
    if not group_by_column:
        return

    group_col_idx = column_index_from_string(group_by_column)
    header_rows = cfg.get("header_rows", 1)
    alt_color = cfg.get("group_colors", {}).get("alternating_fill")
    if not alt_color:
        return

    previous_group_value = None
    group_index = 0
    for row_idx in range(header_rows + 1, sheet.max_row + 1):
        current_value = sheet.cell(row=row_idx, column=group_col_idx).value
        if current_value != previous_group_value:
            group_index += 1
            previous_group_value = current_value
        if group_index % 2 == 0:
            fill_row(sheet, row_idx, alt_color)


def finalize_sheet_style(
    sheet: Worksheet,
    config: dict[str, Any],
    sheet_name: str,
    pre_transforms_sheet: str,
    joins_sheet: str,
    mappings_sheet: str,
) -> None:
    cfg = sheet_config(config, sheet_name)
    apply_base_style(sheet, cfg)
    apply_header_style(sheet, cfg)
    apply_merge_ranges(sheet, cfg)
    center_merged_headers(sheet, cfg)
    apply_special_header_colors(sheet, sheet_name, cfg, pre_transforms_sheet, joins_sheet, mappings_sheet)
    apply_row_heights(sheet, cfg)
    apply_column_widths(sheet, cfg)
    apply_freeze_panes(sheet, cfg)
    apply_autofilter(sheet, cfg)
    apply_alternating_group_fill(sheet, cfg)
    if cfg.get("apply_global_borders", True):
        apply_cell_borders(sheet)


def append_csv_sheet(
    wb: Workbook,
    title: str,
    csv_path: Path,
    config: dict[str, Any],
    required: bool,
    pre_transforms_sheet: str,
    joins_sheet: str,
    mappings_sheet: str,
) -> None:
    if not csv_path.exists():
        if required:
            raise ValueError(f"Required CSV not found: {csv_path}")
        return

    sheet = create_sheet(wb, title)
    headers, rows = read_csv_rows(csv_path)
    if headers:
        sheet.append([repo_header_to_excel(h) for h in headers])
    for row in rows:
        sheet.append(row)

    finalize_sheet_style(sheet, config, title, pre_transforms_sheet, joins_sheet, mappings_sheet)
