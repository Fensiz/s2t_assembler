from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from s2t_tool.infrastructure.common import (
    JOIN_SOURCE_SQL,
    load_json_resource,
    normalize_newlines,
    read_csv_rows,
    read_json_file,
    read_text_file,
)


def build_pre_transforms_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    sheet_name: str,
    pre_transforms_dir: str,
    diff_repo_dir: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    maybe_build_rich_diff: Callable[..., Any],
    pre_transform_row_key: Callable[[str], str],
) -> None:
    diff_enabled = diff_repo_dir is not None

    sheet = create_sheet(wb, sheet_name)
    sheet.cell(row=1, column=1, value="TARGET")
    sheet.cell(row=1, column=3, value="SOURCE")

    headers = [
        "Target table",
        "Source tables",
        "Transformation SQL",
        "Comment",
        "Settings SQL",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    root = repo_dir / pre_transforms_dir
    if not root.exists():
        finalize_sheet_style(sheet, config, sheet_name)
        return

    old_rows_by_key: dict[str, dict[str, str]] = {}
    if diff_enabled:
        old_root = Path(diff_repo_dir) / pre_transforms_dir  # type: ignore[arg-type]
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                json_path = old_table_dir / "pre-transform.json"
                if not json_path.exists():
                    continue

                payload = read_json_file(json_path, default={}) or {}
                target_table = payload.get("target_table", old_table_dir.name)
                old_rows_by_key[pre_transform_row_key(target_table)] = {
                    "source_tables": "\n".join(payload.get("source_tables", [])),
                    "comments": payload.get("comments", ""),
                    "transformation_sql": read_text_file(old_table_dir / "preliminary_transformation.sql"),
                    "settings_sql": read_text_file(old_table_dir / "settings.sql"),
                }

    row_idx = 3
    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        json_path = table_dir / "pre-transform.json"
        if not json_path.exists():
            continue

        payload = read_json_file(json_path, default={}) or {}
        target_table = payload.get("target_table", table_dir.name)
        source_tables = "\n".join(payload.get("source_tables", []))
        comments = payload.get("comments", "")
        transformation_sql = read_text_file(table_dir / "preliminary_transformation.sql")
        settings_sql = read_text_file(table_dir / "settings.sql")

        key = pre_transform_row_key(target_table)
        old_row = old_rows_by_key.get(key, {})

        sheet.cell(row=row_idx, column=1, value=target_table)
        sheet.cell(row=row_idx, column=2, value=maybe_build_rich_diff(diff_enabled, old_row.get("source_tables"), source_tables))
        sheet.cell(row=row_idx, column=3, value=maybe_build_rich_diff(diff_enabled, old_row.get("transformation_sql"), transformation_sql))
        sheet.cell(row=row_idx, column=4, value=maybe_build_rich_diff(diff_enabled, old_row.get("comments"), comments))
        sheet.cell(row=row_idx, column=5, value=maybe_build_rich_diff(diff_enabled, old_row.get("settings_sql"), settings_sql))
        row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)


def build_joins_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    sheet_name: str,
    joins_dir: str,
    diff_repo_dir: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    maybe_build_rich_diff: Callable[..., Any],
    join_row_key: Callable[[str, str], tuple[str, str]],
) -> None:
    diff_enabled = diff_repo_dir is not None
    root = repo_dir / joins_dir
    if not root.exists():
        raise ValueError(f"Required directory not found: {root}")

    old_rows_by_key: dict[tuple[str, str], dict[str, str]] = {}
    if diff_enabled:
        old_root = Path(diff_repo_dir) / joins_dir  # type: ignore[arg-type]
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                for old_load_dir in sorted([p for p in old_table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
                    join_json = read_json_file(old_load_dir / "join.json", default={}) or {}
                    key = join_row_key(old_table_dir.name, old_load_dir.name)
                    old_rows_by_key[key] = {
                        "description": join_json.get("description", ""),
                        "table_codes": "\n".join(join_json.get("table_codes", [])),
                        "table_codes_to_track_delta": "\n".join(join_json.get("table_codes_to_track_delta", [])),
                        "source_tables_join_sql": read_text_file(old_load_dir / JOIN_SOURCE_SQL),
                        "load_code_params": "\n".join(join_json.get("load_code_params", [])),
                        "settings_table_join_sql": read_text_file(old_load_dir / "settings_table_join.sql"),
                        "history_rule": join_json.get("history_rule", ""),
                        "business_history_dates": join_json.get("business_history_dates", ""),
                    }

    sheet = create_sheet(wb, sheet_name)
    sheet["B1"] = "TARGET"
    sheet["F1"] = "SOURCE"
    headers = [
        "Load code",
        "Table name",
        "Description",
        "Table code(s)",
        "Table codes to track delta",
        "Source tables join",
        "Load code params",
        "Settings table join",
        "History rule",
        "Business history dates",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    row_idx = 3
    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        for load_dir in sorted([p for p in table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
            join_json = read_json_file(load_dir / "join.json", default={}) or {}
            key = join_row_key(table_dir.name, load_dir.name)
            old_row = old_rows_by_key.get(key, {})

            current_description = join_json.get("description", "")
            current_table_codes = "\n".join(join_json.get("table_codes", []))
            current_delta_codes = "\n".join(join_json.get("table_codes_to_track_delta", []))
            current_source_join_sql = read_text_file(load_dir / JOIN_SOURCE_SQL)
            current_load_code_params = "\n".join(join_json.get("load_code_params", []))
            current_settings_join_sql = read_text_file(load_dir / "settings_table_join.sql")
            current_history_rule = join_json.get("history_rule", "")
            current_business_history_dates = join_json.get("business_history_dates", "")

            sheet.cell(row=row_idx, column=1, value=load_dir.name)
            sheet.cell(row=row_idx, column=2, value=table_dir.name)
            sheet.cell(row=row_idx, column=3, value=maybe_build_rich_diff(diff_enabled, old_row.get("description"), current_description))
            sheet.cell(row=row_idx, column=4, value=maybe_build_rich_diff(diff_enabled, old_row.get("table_codes"), current_table_codes))
            sheet.cell(row=row_idx, column=5, value=maybe_build_rich_diff(diff_enabled, old_row.get("table_codes_to_track_delta"), current_delta_codes))
            sheet.cell(row=row_idx, column=6, value=maybe_build_rich_diff(diff_enabled, old_row.get("source_tables_join_sql"), current_source_join_sql))
            sheet.cell(row=row_idx, column=7, value=maybe_build_rich_diff(diff_enabled, old_row.get("load_code_params"), current_load_code_params))
            sheet.cell(row=row_idx, column=8, value=maybe_build_rich_diff(diff_enabled, old_row.get("settings_table_join_sql"), current_settings_join_sql))
            sheet.cell(row=row_idx, column=9, value=maybe_build_rich_diff(diff_enabled, old_row.get("history_rule"), current_history_rule))
            sheet.cell(row=row_idx, column=10, value=maybe_build_rich_diff(diff_enabled, old_row.get("business_history_dates"), current_business_history_dates))
            row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)


def resolve_attribute_name(table_name: str, attribute_code: str, attribute_names: dict[str, Any]) -> str:
    tables = attribute_names.get("tables", {}) or {}
    common = attribute_names.get("common", {}) or {}
    if table_name in tables and attribute_code in tables[table_name]:
        return tables[table_name][attribute_code]
    return common.get(attribute_code, "")


def build_mappings_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    sheet_name: str,
    joins_dir: str,
    attribute_names_json: str,
    mappings_csv: str,
    mappings_extra_json: str,
    diff_repo_dir: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    maybe_build_rich_diff: Callable[..., Any],
    mapping_row_key: Callable[[str, str, str], tuple[str, str, str]],
) -> None:
    diff_enabled = diff_repo_dir is not None
    old_rows_by_key: dict[tuple[str, str, str], dict[str, str]] = {}

    if diff_enabled:
        old_root = Path(diff_repo_dir) / joins_dir  # type: ignore[arg-type]
        old_attribute_names = read_json_file(Path(diff_repo_dir) / attribute_names_json, default={}) or {}
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                for old_load_dir in sorted([p for p in old_table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
                    mappings_path = old_load_dir / mappings_csv
                    if not mappings_path.exists():
                        continue

                    _, old_rows = read_csv_rows(mappings_path)
                    old_extra = read_json_file(old_load_dir / mappings_extra_json, default={}) or {}
                    for row in old_rows:
                        while len(row) < 2:
                            row.append("")
                        attribute_code = row[0]
                        extra_item = old_extra.get(attribute_code, {}) or {}
                        key = mapping_row_key(old_load_dir.name, old_table_dir.name, attribute_code)
                        old_rows_by_key[key] = {
                            "attribute_name": resolve_attribute_name(old_table_dir.name, attribute_code, old_attribute_names),
                            "mapping_algorithm": row[1],
                            "additional_join": extra_item.get("additional_join", ""),
                            "settings": extra_item.get("settings", ""),
                        }

    root = repo_dir / joins_dir
    attribute_names = read_json_file(repo_dir / attribute_names_json, default={}) or {}
    sheet = create_sheet(wb, sheet_name)
    sheet.cell(row=1, column=2, value="TARGET")
    sheet.cell(row=1, column=5, value="SOURCE")

    headers = [
        "Load code",
        "Table name",
        "Attribute code",
        "Attribute name",
        "Mapping algorithm",
        "Additional join",
        "Settings",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    row_idx = 3
    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        for load_dir in sorted([p for p in table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
            mappings_path = load_dir / mappings_csv
            if not mappings_path.exists():
                continue

            _, rows = read_csv_rows(mappings_path)
            extra = read_json_file(load_dir / mappings_extra_json, default={}) or {}
            for row in rows:
                while len(row) < 2:
                    row.append("")
                attribute_code = row[0]
                mapping_algorithm = normalize_newlines(row[1])
                extra_item = extra.get(attribute_code, {}) or {}

                sheet.cell(row=row_idx, column=1, value=load_dir.name)
                sheet.cell(row=row_idx, column=2, value=table_dir.name)
                sheet.cell(row=row_idx, column=3, value=attribute_code)

                key = mapping_row_key(load_dir.name, table_dir.name, attribute_code)
                old_row = old_rows_by_key.get(key, {})
                current_attribute_name = resolve_attribute_name(table_dir.name, attribute_code, attribute_names)
                current_additional_join = extra_item.get("additional_join", "")
                current_settings = extra_item.get("settings", "")

                sheet.cell(row=row_idx, column=4, value=maybe_build_rich_diff(diff_enabled, old_row.get("attribute_name"), current_attribute_name))
                sheet.cell(row=row_idx, column=5, value=maybe_build_rich_diff(diff_enabled, old_row.get("mapping_algorithm"), mapping_algorithm))
                sheet.cell(row=row_idx, column=6, value=maybe_build_rich_diff(diff_enabled, old_row.get("additional_join"), current_additional_join))
                sheet.cell(row=row_idx, column=7, value=maybe_build_rich_diff(diff_enabled, old_row.get("settings"), current_settings))
                row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)


def normalize_table_row(values: Any, headers: list[str] | None) -> list[Any]:
    if isinstance(values, list):
        return values
    if isinstance(values, dict):
        if headers:
            return [values.get(header, "") for header in headers]
        return list(values.values())
    return [values]


def build_metadata_sheet(
    wb: Workbook,
    config: dict[str, Any],
    sheet_name: str,
    metadata_json: str,
    diff_commit: str | None,
    create_sheet: Callable[..., Any],
    finalize_sheet_style: Callable[..., None],
    apply_table_borders: Callable[..., None],
) -> None:
    payload = load_json_resource(metadata_json) or {}
    blocks = payload.get("blocks", [])
    sheet = create_sheet(wb, sheet_name)
    row_idx = 1

    if diff_commit:
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        cell = sheet.cell(row=row_idx, column=1, value=f"DIFF FILE — DO NOT USE FOR PUT (compared with commit: {diff_commit})")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_idx += 2

    for block in blocks:
        title_en = block.get("title_en")
        title_ru = block.get("title_ru")
        link = block.get("link")
        table = block.get("table")

        if title_en:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=title_en)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

        if title_ru:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=title_ru)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

        if link:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=link)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        for text_line in block.get("lines", []) or []:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=text_line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        if table:
            headers = table.get("headers")
            rows = table.get("rows", []) or []
            table_start_row = row_idx
            header_names: list[str] | None = None

            if headers:
                header_names = [str(h) for h in headers]
                for col_idx, header in enumerate(header_names, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row_idx += 1

            max_width = len(header_names) if header_names else 0
            for raw_row in rows:
                normalized_row = normalize_table_row(raw_row, header_names)
                max_width = max(max_width, len(normalized_row))
                for col_idx, value in enumerate(normalized_row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                row_idx += 1

            table_end_row = row_idx - 1
            if max_width > 0 and table_end_row >= table_start_row:
                apply_table_borders(sheet, table_start_row, table_end_row, 1, max_width)

        row_idx += 1

    finalize_sheet_style(sheet, config, sheet_name)
