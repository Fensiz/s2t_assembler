from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from s2t_tool.domain.schema import DEFAULT_SCHEMA, S2TSchema
from s2t_tool.shared.csv_files import write_csv_rows
from s2t_tool.shared.files import ensure_dir, write_json_file, write_text_file
from s2t_tool.shared.text import (
    excel_to_repo_header,
    is_row_empty,
    normalize_cell,
    slugify_dir_name,
    split_lines,
)


def sheet_rows(sheet: Worksheet) -> list[tuple[Any, ...]]:
    return list(sheet.iter_rows(values_only=True))


class ExcelRepoReader:
    def __init__(
        self,
        excel_path: str | Path,
        output_dir: str | Path,
        schema: S2TSchema = DEFAULT_SCHEMA,
        logger=None,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.schema = schema
        self.logger = logger
        self.workbook: Workbook = load_workbook(self.excel_path, data_only=True)

    def get_sheet(self, sheet_name: str, required: bool = True) -> Worksheet | None:
        normalized_aliases = {
            self.schema.normalize_sheet_name(alias)
            for alias in self.schema.sheet_aliases_for(sheet_name)
        }

        for actual_name in self.workbook.sheetnames:
            if self.schema.normalize_sheet_name(actual_name) in normalized_aliases:
                return self.workbook[actual_name]

        if required:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {self.excel_path}. "
                f"Available sheets: {self.workbook.sheetnames}"
            )

        return None

    def export_change_history(self) -> Path:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(self.schema.change_history_sheet, required=True)
        assert sheet is not None
        rows = sheet_rows(sheet)

        entries: list[dict[str, str | None]] = []
        for row in rows[1:]:
            values = [normalize_cell(v) for v in row[:5]]
            while len(values) < 5:
                values.append("")
            if is_row_empty(values):
                continue

            entries.append(
                {
                    "author": values[0] or None,
                    "date": values[1] or None,
                    "version": values[2] or None,
                    "description": values[3] or None,
                    "jira_ticket": values[4] or None,
                }
            )

        path = self.output_dir / self.schema.change_history_json
        write_json_file(path, entries)
        return path

    def export_source_lg(self) -> Path:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(self.schema.source_lg_sheet, required=True)
        assert sheet is not None
        rows = sheet_rows(sheet)
        if not rows:
            raise ValueError("Source LG sheet is empty")

        header_map: dict[int, str] = {}
        source_headers = list(self.schema.source_lg_headers)
        for idx, name in enumerate(rows[0]):
            normalized = excel_to_repo_header(str(name or ""))
            if normalized in source_headers:
                header_map[idx] = normalized

        missing = set(source_headers) - set(header_map.values())
        if missing:
            raise ValueError(f"Missing columns in Source LG: {sorted(missing)}")

        output_rows: list[list[str]] = []
        for row in rows[1:]:
            values = {name: "" for name in source_headers}
            for idx, repo_name in header_map.items():
                if idx < len(row):
                    values[repo_name] = normalize_cell(row[idx])
            if is_row_empty(list(values.values())):
                continue
            output_rows.append([values[h] for h in source_headers])

        path = self.output_dir / self.schema.source_lg_csv
        write_csv_rows(path, source_headers, output_rows)
        return path

    def export_targets(self) -> Path:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(self.schema.targets_sheet, required=True)
        assert sheet is not None
        rows = sheet_rows(sheet)

        output_rows: list[list[str]] = []
        for row in rows[1:]:
            values = [normalize_cell(v) for v in row[:2]]
            while len(values) < 2:
                values.append("")
            if is_row_empty(values):
                continue
            output_rows.append(values)

        path = self.output_dir / self.schema.targets_csv
        write_csv_rows(path, list(self.schema.targets_headers), output_rows)
        return path

    def export_simple_csv_sheet(
        self,
        sheet_name: str,
        output_name: str,
        headers: list[str],
        required: bool = False,
    ) -> Path | None:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(sheet_name, required=required)
        if sheet is None:
            return None

        rows = sheet_rows(sheet)
        output_rows: list[list[str]] = []

        for row in rows[1:]:
            values = [normalize_cell(v) for v in row[:len(headers)]]
            while len(values) < len(headers):
                values.append("")
            if is_row_empty(values):
                continue
            output_rows.append(values)

        path = self.output_dir / output_name
        write_csv_rows(path, headers, output_rows)
        return path

    def export_pre_transforms(self) -> Path:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(self.schema.pre_transforms_sheet, required=True)
        assert sheet is not None
        rows = sheet_rows(sheet)

        if len(rows) < 2:
            raise ValueError("Pre-transforms sheet must contain at least double header")

        root = self.output_dir / self.schema.pre_transforms_dir
        ensure_dir(root)

        if len(rows) == 2:
            return root

        for row in rows[2:]:
            values = [normalize_cell(v) for v in row[:5]]
            while len(values) < 5:
                values.append("")
            if is_row_empty(values):
                continue

            target_table, source_tables_raw, transformation_sql, comments, settings_sql = values
            if not target_table:
                raise ValueError("Pre-transforms row has empty target table")

            target_dir = root / slugify_dir_name(target_table)
            ensure_dir(target_dir)

            payload = {
                "target_table": target_table,
                "source_tables": split_lines(source_tables_raw),
                "comments": comments or None,
            }

            write_json_file(target_dir / "pre-transform.json", payload)

            if transformation_sql:
                write_text_file(target_dir / "preliminary_transformation.sql", transformation_sql)
            if settings_sql:
                write_text_file(target_dir / "settings.sql", settings_sql)

        return root

    def export_joins(self) -> Path:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(self.schema.joins_sheet, required=True)
        assert sheet is not None
        rows = sheet_rows(sheet)
        if len(rows) < 3:
            raise ValueError("Joins must contain double header and data rows")

        root = self.output_dir / self.schema.joins_dir
        ensure_dir(root)

        for row in rows[2:]:
            values = [normalize_cell(v) for v in row[:10]]
            while len(values) < 10:
                values.append("")
            if is_row_empty(values):
                continue

            (
                load_code,
                table_name,
                description,
                table_codes_raw,
                delta_codes_raw,
                source_tables_join_sql,
                load_code_params_raw,
                settings_table_join_sql,
                history_rule,
                business_history_dates,
            ) = values

            if not load_code or not table_name:
                raise ValueError("Joins row must contain table_name and load_code")

            join_dir = root / slugify_dir_name(table_name) / slugify_dir_name(load_code)
            ensure_dir(join_dir)

            payload = {
                "description": description or None,
                "table_codes": split_lines(table_codes_raw),
                "table_codes_to_track_delta": split_lines(delta_codes_raw),
                "load_code_params": split_lines(load_code_params_raw),
                "history_rule": history_rule or None,
                "business_history_dates": business_history_dates or None,
            }
            payload = {k: v for k, v in payload.items() if v not in (None, [], "")}
            write_json_file(join_dir / "join.json", payload)

            if source_tables_join_sql:
                write_text_file(join_dir / "source_tables_join.sql", source_tables_join_sql)
            if settings_table_join_sql:
                write_text_file(join_dir / "settings_table_join.sql", settings_table_join_sql)

        return root

    def export_mappings(self) -> tuple[Path, Path]:
        ensure_dir(self.output_dir)

        sheet = self.get_sheet(self.schema.mappings_sheet, required=True)
        assert sheet is not None
        rows = sheet_rows(sheet)
        if len(rows) < 2:
            raise ValueError("Mappings sheet must contain header and data rows")

        joins_root = self.output_dir / self.schema.joins_dir
        ensure_dir(joins_root)

        grouped_rows: dict[tuple[str, str], list[dict[str, str]]] = {}
        attribute_names_by_table: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))

        for row in rows[2:]:
            values = [normalize_cell(v) for v in row[:7]]
            while len(values) < 7:
                values.append("")
            if is_row_empty(values):
                continue

            load_code, table_name, attribute_code, attribute_name, mapping_algorithm, additional_join, settings = values

            if not load_code or not table_name or not attribute_code:
                raise ValueError("Mappings row must contain load_code, table_name, attribute_code")

            grouped_rows.setdefault((table_name, load_code), []).append(
                {
                    "attribute_code": attribute_code,
                    "mapping_algorithm": mapping_algorithm,
                    "additional_join": additional_join,
                    "settings": settings,
                }
            )

            if attribute_name:
                attribute_names_by_table[table_name][attribute_code].append(attribute_name)

        for (table_name, load_code), mapping_rows in grouped_rows.items():
            mapping_dir = joins_root / slugify_dir_name(table_name) / slugify_dir_name(load_code)
            ensure_dir(mapping_dir)

            csv_rows: list[list[str]] = []
            extra_payload: dict[str, dict[str, str]] = {}

            for item in mapping_rows:
                csv_rows.append([item["attribute_code"], item["mapping_algorithm"]])

                extra: dict[str, str] = {}
                if item["additional_join"]:
                    extra["additional_join"] = item["additional_join"]
                if item["settings"]:
                    extra["settings"] = item["settings"]

                if extra:
                    extra_payload[item["attribute_code"]] = extra

            write_csv_rows(
                mapping_dir / self.schema.mappings_csv,
                ["attribute_code", "mapping_algorithm"],
                csv_rows,
            )

            if extra_payload:
                write_json_file(mapping_dir / self.schema.mappings_extra_json, extra_payload)

        resolved_table_names: dict[str, dict[str, str]] = {}
        for table_name, attrs in attribute_names_by_table.items():
            resolved_table_names[table_name] = {}
            for attribute_code, names in attrs.items():
                unique_names = {n for n in names if n}
                if len(unique_names) > 1:
                    raise ValueError(
                        f"Conflicting attribute_name inside table '{table_name}' "
                        f"for attribute_code '{attribute_code}': {sorted(unique_names)}"
                    )
                if unique_names:
                    resolved_table_names[table_name][attribute_code] = next(iter(unique_names))

        attribute_to_table_values: dict[str, dict[str, str]] = defaultdict(dict)
        for table_name, attrs in resolved_table_names.items():
            for attribute_code, attribute_name in attrs.items():
                attribute_to_table_values[attribute_code][table_name] = attribute_name

        common_names: dict[str, str] = {}
        table_overrides: dict[str, dict[str, str]] = {}

        for attribute_code, table_values in attribute_to_table_values.items():
            unique_values = set(table_values.values())
            if len(unique_values) == 1:
                common_names[attribute_code] = next(iter(unique_values))
            else:
                for table_name, attribute_name in table_values.items():
                    table_overrides.setdefault(table_name, {})[attribute_code] = attribute_name

        attribute_names_payload = {
            "common": dict(sorted(common_names.items())),
            "tables": {
                table_name: dict(sorted(attrs.items()))
                for table_name, attrs in sorted(table_overrides.items())
            },
        }

        attribute_names_path = self.output_dir / self.schema.attribute_names_json
        write_json_file(attribute_names_path, attribute_names_payload)
        return joins_root, attribute_names_path

    def export_all(self) -> None:
        created: list[Path] = [
            self.export_change_history(),
            self.export_source_lg(),
            self.export_targets(),
            self.export_pre_transforms(),
            self.export_joins(),
        ]

        mappings_root, attribute_names_path = self.export_mappings()
        created.append(mappings_root)
        created.append(attribute_names_path)

        for maybe_path in [
            self.export_simple_csv_sheet(
                self.schema.settings_sheet,
                self.schema.settings_csv,
                list(self.schema.settings_headers),
                required=False,
            ),
            self.export_simple_csv_sheet(
                self.schema.parameters_sheet,
                self.schema.parameters_csv,
                list(self.schema.parameters_headers),
                required=False,
            ),
            self.export_simple_csv_sheet(
                self.schema.st_decoder_sheet,
                self.schema.st_decoder_csv,
                list(self.schema.settings_headers),
                required=False,
            ),
            self.export_simple_csv_sheet(
                self.schema.st_filter_sheet,
                self.schema.st_filter_csv,
                list(self.schema.st_filter_headers),
                required=False,
            ),
        ]:
            if maybe_path is not None:
                created.append(maybe_path)

        for path in created:
            if self.logger:
                self.logger(f"Created: {path}")


def normalize_sheet_name(value: str) -> str:
    return DEFAULT_SCHEMA.normalize_sheet_name(value)


def get_sheet(excel_path: Path, sheet_name: str, required: bool = True) -> Worksheet | None:
    reader = ExcelRepoReader(excel_path=excel_path, output_dir=excel_path.parent)
    return reader.get_sheet(sheet_name, required=required)


def export_change_history(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_change_history()


def export_source_lg(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_source_lg()


def export_targets(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_targets()


def export_simple_csv_sheet(
    excel_path: Path,
    output_dir: Path,
    sheet_name: str,
    output_name: str,
    headers: list[str],
    required: bool = False,
) -> Path | None:
    return ExcelRepoReader(excel_path, output_dir).export_simple_csv_sheet(
        sheet_name=sheet_name,
        output_name=output_name,
        headers=headers,
        required=required,
    )


def export_pre_transforms(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_pre_transforms()


def export_joins(excel_path: Path, output_dir: Path) -> Path:
    return ExcelRepoReader(excel_path, output_dir).export_joins()


def export_mappings(excel_path: Path, output_dir: Path) -> tuple[Path, Path]:
    return ExcelRepoReader(excel_path, output_dir).export_mappings()


def section_key_from_title(title: str) -> str:
    return (
        title.strip()
        .lower()
        .replace("-", " ")
        .replace("/", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace(":", "")
        .replace(".", "")
        .replace(",", "")
        .replace("  ", " ")
        .replace(" ", "_")
    )


def export_excel_to_repo(excel_path: str, output_dir: str, logger=None) -> None:
    ExcelRepoReader(
        excel_path=excel_path,
        output_dir=output_dir,
        schema=DEFAULT_SCHEMA,
        logger=logger,
    ).export_all()


if __name__ == "__main__":
    export_excel_to_repo("input.xlsx", "./repo")
