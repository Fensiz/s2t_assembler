from __future__ import annotations

from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from s2t_tool.domain.schema import DEFAULT_SCHEMA, S2TSchema
from s2t_tool.infrastructure.writer_style import (
    WRITER_CONFIG_FILE,
    apply_table_borders,
    append_csv_sheet,
    create_sheet,
    finalize_sheet_style,
    load_writer_config,
)
from s2t_tool.infrastructure.writer_sections import (
    build_joins_sheet as build_joins_sheet_section,
    build_mappings_sheet as build_mappings_sheet_section,
    build_metadata_sheet as build_metadata_sheet_section,
    build_pre_transforms_sheet as build_pre_transforms_sheet_section,
    resolve_attribute_name,
)
from s2t_tool.shared.files import read_json_file


SCHEMA = DEFAULT_SCHEMA

CHANGE_HISTORY_JSON = SCHEMA.change_history_json
SOURCE_LG_CSV = SCHEMA.source_lg_csv
TARGETS_CSV = SCHEMA.targets_csv
SETTINGS_CSV = SCHEMA.settings_csv
PARAMETERS_CSV = SCHEMA.parameters_csv
ST_DECODER_CSV = SCHEMA.st_decoder_csv
ST_FILTER_CSV = SCHEMA.st_filter_csv
METADATA_JSON = SCHEMA.metadata_json
ATTRIBUTE_NAMES_JSON = SCHEMA.attribute_names_json

PRE_TRANSFORMS_DIR = SCHEMA.pre_transforms_dir
JOINS_DIR = SCHEMA.joins_dir

CHANGE_HISTORY_SHEET = SCHEMA.change_history_sheet
SOURCE_LG_SHEET = SCHEMA.source_lg_sheet
TARGETS_SHEET = SCHEMA.targets_sheet
PRE_TRANSFORMS_SHEET = SCHEMA.pre_transforms_sheet
JOINS_SHEET = SCHEMA.joins_sheet
MAPPINGS_SHEET = SCHEMA.mappings_sheet
SETTINGS_SHEET = SCHEMA.settings_sheet
PARAMETERS_SHEET = SCHEMA.parameters_sheet
ST_DECODER_SHEET = SCHEMA.st_decoder_sheet
ST_FILTER_SHEET = SCHEMA.st_filter_sheet
METADATA_SHEET = SCHEMA.metadata_sheet

BLACK_FONT = InlineFont(color="000000")
GREEN_FONT = InlineFont(color="008000")
RED_FONT = InlineFont(color="FF0000", strike=True)


# ============================================================
# Diff helpers
# ============================================================

def build_rich_diff(old_text: str | None, new_text: str | None) -> CellRichText | str:
    old_value = "" if old_text is None else str(old_text)
    new_value = "" if new_text is None else str(new_text)

    if old_value == new_value:
        return new_value

    matcher = SequenceMatcher(a=old_value, b=new_value)
    rich = CellRichText()

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        old_chunk = old_value[i1:i2]
        new_chunk = new_value[j1:j2]

        if tag == "equal":
            if old_chunk:
                rich.append(old_chunk)
        elif tag == "delete":
            if old_chunk:
                rich.append(TextBlock(RED_FONT, old_chunk))
        elif tag == "insert":
            if new_chunk:
                rich.append(TextBlock(GREEN_FONT, new_chunk))
        elif tag == "replace":
            if old_chunk:
                rich.append(TextBlock(RED_FONT, old_chunk))
            if new_chunk:
                rich.append(TextBlock(GREEN_FONT, new_chunk))

    return rich


def maybe_build_rich_diff(
    diff_enabled: bool,
    old_value: str | None,
    new_value: str | None,
) -> str | CellRichText:
    """
    Return plain text in normal mode and rich diff in diff mode.
    """
    new_text = "" if new_value is None else str(new_value)

    if not diff_enabled:
        return new_text

    return build_rich_diff(old_value, new_text)


def normalize_key_part(value: object) -> str:
    return "" if value is None else str(value)


def join_row_key(table_name: str, load_code: str) -> tuple[str, str]:
    return (
        normalize_key_part(table_name),
        normalize_key_part(load_code),
    )


def pre_transform_row_key(target_table: str) -> str:
    return normalize_key_part(target_table)


def mapping_row_key(load_code: str, table_name: str, attribute_code: str) -> tuple[str, str, str]:
    return (
        normalize_key_part(load_code),
        normalize_key_part(table_name),
        normalize_key_part(attribute_code),
    )


def build_change_history_sheet(wb: Workbook, repo_dir: Path, config: dict[str, Any]) -> None:
    path = repo_dir / CHANGE_HISTORY_JSON
    entries = read_json_file(path, default=[])
    sheet = create_sheet(wb, CHANGE_HISTORY_SHEET)

    sheet.append(["Author", "Date", "Version", "Description", "Jira ticket"])

    for entry in entries:
        sheet.append(
            [
                entry.get("author"),
                entry.get("date"),
                entry.get("version"),
                entry.get("description"),
                entry.get("jira_ticket"),
            ]
        )

    finalize_sheet_style(sheet, config, CHANGE_HISTORY_SHEET, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)


def build_source_lg_sheet(wb: Workbook, repo_dir: Path, config: dict[str, Any]) -> None:
    append_csv_sheet(wb, SOURCE_LG_SHEET, repo_dir / SOURCE_LG_CSV, config, True, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)


def build_targets_sheet(wb: Workbook, repo_dir: Path, config: dict[str, Any]) -> None:
    append_csv_sheet(wb, TARGETS_SHEET, repo_dir / TARGETS_CSV, config, True, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)


# ============================================================
# Pre-transforms
# ============================================================

def build_pre_transforms_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_pre_transforms_sheet_section(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        sheet_name=PRE_TRANSFORMS_SHEET,
        pre_transforms_dir=PRE_TRANSFORMS_DIR,
        diff_repo_dir=diff_repo_dir,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        maybe_build_rich_diff=maybe_build_rich_diff,
        pre_transform_row_key=pre_transform_row_key,
    )


# ============================================================
# Joins
# ============================================================

def build_joins_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_joins_sheet_section(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        sheet_name=JOINS_SHEET,
        joins_dir=JOINS_DIR,
        diff_repo_dir=diff_repo_dir,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        maybe_build_rich_diff=maybe_build_rich_diff,
        join_row_key=join_row_key,
    )


# ============================================================
# Mappings
# ============================================================

def build_mappings_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    build_mappings_sheet_section(
        wb=wb,
        repo_dir=repo_dir,
        config=config,
        sheet_name=MAPPINGS_SHEET,
        joins_dir=JOINS_DIR,
        attribute_names_json=ATTRIBUTE_NAMES_JSON,
        mappings_csv="mappings.csv",
        mappings_extra_json="mappings.extra.json",
        diff_repo_dir=diff_repo_dir,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        maybe_build_rich_diff=maybe_build_rich_diff,
        mapping_row_key=mapping_row_key,
    )


# ============================================================
# Metadata
# ============================================================

def build_metadata_sheet(
    wb: Workbook,
    config: dict[str, Any],
    diff_commit: str | None = None,
) -> None:
    build_metadata_sheet_section(
        wb=wb,
        config=config,
        sheet_name=METADATA_SHEET,
        metadata_json=METADATA_JSON,
        diff_commit=diff_commit,
        create_sheet=create_sheet,
        finalize_sheet_style=lambda sheet, cfg, name: finalize_sheet_style(
            sheet, cfg, name, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET
        ),
        apply_table_borders=apply_table_borders,
    )


# ============================================================
# Main entry
# ============================================================

def build_excel_from_repo(
    repo_dir: str,
    output_excel_path: str,
    config_path: str = WRITER_CONFIG_FILE,
    diff_repo_dir: str | None = None,
    diff_commit: str | None = None,
    logger=None,
) -> None:
    RepoExcelWriter(
        repo_dir=repo_dir,
        output_excel_path=output_excel_path,
        config_path=config_path,
        diff_repo_dir=diff_repo_dir,
        diff_commit=diff_commit,
        logger=logger,
        schema=DEFAULT_SCHEMA,
    ).build()


class RepoExcelWriter:
    def __init__(
        self,
        repo_dir: str | Path,
        output_excel_path: str | Path,
        config_path: str | Path = WRITER_CONFIG_FILE,
        diff_repo_dir: str | Path | None = None,
        diff_commit: str | None = None,
        logger=None,
        schema: S2TSchema = DEFAULT_SCHEMA,
    ) -> None:
        self.repo_dir = Path(repo_dir)
        self.output_excel_path = Path(output_excel_path)
        self.config = load_writer_config(config_path)
        self.diff_repo_dir = str(diff_repo_dir) if diff_repo_dir is not None else None
        self.diff_commit = diff_commit
        self.logger = logger
        self.schema = schema

    def build(self) -> None:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        self.build_change_history_sheet(wb)
        self.build_source_lg_sheet(wb)
        self.build_targets_sheet(wb)
        self.build_pre_transforms_sheet(wb)
        self.build_joins_sheet(wb)
        self.build_mappings_sheet(wb)
        self.append_optional_csv_sheets(wb)
        self.build_metadata_sheet(wb)

        wb.save(self.output_excel_path)
        if self.logger:
            self.logger(f"Created: {self.output_excel_path}")

    def build_change_history_sheet(self, wb: Workbook) -> None:
        build_change_history_sheet(wb, self.repo_dir, self.config)

    def build_source_lg_sheet(self, wb: Workbook) -> None:
        build_source_lg_sheet(wb, self.repo_dir, self.config)

    def build_targets_sheet(self, wb: Workbook) -> None:
        build_targets_sheet(wb, self.repo_dir, self.config)

    def build_pre_transforms_sheet(self, wb: Workbook) -> None:
        build_pre_transforms_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_joins_sheet(self, wb: Workbook) -> None:
        build_joins_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def build_mappings_sheet(self, wb: Workbook) -> None:
        build_mappings_sheet(wb, self.repo_dir, self.config, self.diff_repo_dir)

    def append_optional_csv_sheets(self, wb: Workbook) -> None:
        append_csv_sheet(wb, self.schema.settings_sheet, self.repo_dir / self.schema.settings_csv, self.config, False, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)
        append_csv_sheet(wb, self.schema.parameters_sheet, self.repo_dir / self.schema.parameters_csv, self.config, False, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)
        append_csv_sheet(wb, self.schema.st_decoder_sheet, self.repo_dir / self.schema.st_decoder_csv, self.config, False, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)
        append_csv_sheet(wb, self.schema.st_filter_sheet, self.repo_dir / self.schema.st_filter_csv, self.config, False, PRE_TRANSFORMS_SHEET, JOINS_SHEET, MAPPINGS_SHEET)

    def build_metadata_sheet(self, wb: Workbook) -> None:
        build_metadata_sheet(wb, self.config, self.diff_commit)


if __name__ == "__main__":
    build_excel_from_repo("./repo", "output.xlsx", WRITER_CONFIG_FILE)
