from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from common import (
    ensure_dir,
    excel_to_repo_header,
    is_row_empty,
    normalize_cell,
    read_json_file,
    slugify_dir_name,
    split_lines,
    write_csv_rows,
    write_json_file,
    write_text_file,
)

CHANGE_HISTORY_SHEET = "Change history"
SOURCE_LG_SHEET = "Source LG"
TARGETS_SHEET = "Targets"
PRE_TRANSFORMS_SHEET = "Pre-transforms"
JOINS_SHEET = "Joins"
MAPPINGS_SHEET = "Mappings"
SETTINGS_SHEET = "Settings"
PARAMETERS_SHEET = "Parameters"
ST_DECODER_SHEET = "ST_DECODER"
ST_FILTER_SHEET = "ST_FILTER"
METADATA_SHEET = "Metadata"

CHANGE_HISTORY_JSON = "change-history.json"
SOURCE_LG_CSV = "source-lg.csv"
TARGETS_CSV = "targets.csv"
SETTINGS_CSV = "settings.csv"
PARAMETERS_CSV = "parameters.csv"
ST_DECODER_CSV = "st_decoder.csv"
ST_FILTER_CSV = "st_filter.csv"
METADATA_JSON = "metadata.json"
ATTRIBUTE_NAMES_JSON = "attribute_names.json"
MAPPINGS_CSV = "mappings.csv"
MAPPINGS_EXTRA_JSON = "mappings.extra.json"

PRE_TRANSFORMS_DIR = "pre-transforms"
JOINS_DIR = "joins"

SOURCE_LG_HEADERS = [
    "scheme",
    "table",
    "column",
    "data_type",
    "data_length",
    "is_key",
    "description",
    "link",
]
TARGETS_HEADERS = ["table_code", "table_name"]
SETTINGS_HEADERS = [
    "settings_alias",
    "settings_description",
    "settings_table",
    "settings_type",
    "period",
    "mask",
]
PARAMETERS_HEADERS = ["parameter", "value", "data_type", "comment"]
ST_DECODER_HEADERS = SETTINGS_HEADERS[:]
ST_FILTER_HEADERS = [
    "settings_type",
    "source_value",
    "target_value",
    "start_dt",
    "end_dt",
    "update_date",
    "author_update_name",
]

SHEET_ALIASES: dict[str, list[str]] = {
    CHANGE_HISTORY_SHEET: [
        "Change history",
        "Change History",
        "ChangeHistory",
        "Change history ",
    ],
    SOURCE_LG_SHEET: [
        "Source LG",
        "SourceLG",
        "Sources LG",
        "SourcesLG",
    ],
    TARGETS_SHEET: [
        "Targets",
        "Target",
    ],
    PRE_TRANSFORMS_SHEET: [
        "Pre-transforms",
        "Pre transforms",
        "PreTransforms",
        "Pre-Transforms",
    ],
    JOINS_SHEET: [
        "Joins",
        "Join",
    ],
    MAPPINGS_SHEET: [
        "Mappings",
        "Mapping",
    ],
    SETTINGS_SHEET: [
        "Settings",
        "Setting",
    ],
    PARAMETERS_SHEET: [
        "Parameters",
        "Parameter",
    ],
    ST_DECODER_SHEET: [
        "ST_DECODER",
        "ST DECODER",
        "ST-DECODER",
        "Decoder",
    ],
    ST_FILTER_SHEET: [
        "ST_FILTER",
        "ST FILTER",
        "ST-FILTER",
        "Filter",
    ],
    METADATA_SHEET: [
        "Metadata",
        "MetaData",
        "Meta Data",
    ],
}

def normalize_sheet_name(value: str) -> str:
    return "".join(str(value).strip().lower().split())


def get_sheet(excel_path: Path, sheet_name: str, required: bool = True) -> Worksheet | None:
    wb = load_workbook(excel_path, data_only=True)

    aliases = SHEET_ALIASES.get(sheet_name, [sheet_name])
    normalized_aliases = {normalize_sheet_name(alias) for alias in aliases}

    for actual_name in wb.sheetnames:
        if normalize_sheet_name(actual_name) in normalized_aliases:
            return wb[actual_name]

    if required:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {excel_path}. "
            f"Available sheets: {wb.sheetnames}"
        )

    return None


def sheet_rows(sheet: Worksheet) -> list[tuple[Any, ...]]:
    return list(sheet.iter_rows(values_only=True))


def export_change_history(excel_path: Path, output_dir: Path) -> Path:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, CHANGE_HISTORY_SHEET, required=True)
    assert sheet is not None
    rows = sheet_rows(sheet)

    entries: list[dict[str, str | None]] = []
    for row in rows[1:]:
        values = [normalize_cell(v) for v in row[:5]]
        while len(values) < 5:
            values.append("")
        if is_row_empty(values):
            continue

        entries.append(
            {
                "author": values[0] or None,
                "date": values[1] or None,
                "version": values[2] or None,
                "description": values[3] or None,
                "jira_ticket": values[4] or None,
            }
        )

    path = output_dir / CHANGE_HISTORY_JSON
    write_json_file(path, entries)
    return path


def export_source_lg(excel_path: Path, output_dir: Path) -> Path:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, SOURCE_LG_SHEET, required=True)
    assert sheet is not None
    rows = sheet_rows(sheet)
    if not rows:
        raise ValueError("Source LG sheet is empty")

    header_map: dict[int, str] = {}
    for idx, name in enumerate(rows[0]):
        normalized = excel_to_repo_header(str(name or ""))
        if normalized in SOURCE_LG_HEADERS:
            header_map[idx] = normalized

    missing = set(SOURCE_LG_HEADERS) - set(header_map.values())
    if missing:
        raise ValueError(f"Missing columns in Source LG: {sorted(missing)}")

    output_rows: list[list[str]] = []
    for row in rows[1:]:
        values = {name: "" for name in SOURCE_LG_HEADERS}
        for idx, repo_name in header_map.items():
            if idx < len(row):
                values[repo_name] = normalize_cell(row[idx])
        if is_row_empty(list(values.values())):
            continue
        output_rows.append([values[h] for h in SOURCE_LG_HEADERS])

    path = output_dir / SOURCE_LG_CSV
    write_csv_rows(path, SOURCE_LG_HEADERS, output_rows)
    return path


def export_targets(excel_path: Path, output_dir: Path) -> Path:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, TARGETS_SHEET, required=True)
    assert sheet is not None
    rows = sheet_rows(sheet)

    output_rows: list[list[str]] = []
    for row in rows[1:]:
        values = [normalize_cell(v) for v in row[:2]]
        while len(values) < 2:
            values.append("")
        if is_row_empty(values):
            continue
        output_rows.append(values)

    path = output_dir / TARGETS_CSV
    write_csv_rows(path, TARGETS_HEADERS, output_rows)
    return path


def export_simple_csv_sheet(
    excel_path: Path,
    output_dir: Path,
    sheet_name: str,
    output_name: str,
    headers: list[str],
    required: bool = False,
) -> Path | None:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, sheet_name, required=required)
    if sheet is None:
        return None

    rows = sheet_rows(sheet)
    output_rows: list[list[str]] = []

    for row in rows[1:]:
        values = [normalize_cell(v) for v in row[:len(headers)]]
        while len(values) < len(headers):
            values.append("")
        if is_row_empty(values):
            continue
        output_rows.append(values)

    path = output_dir / output_name
    write_csv_rows(path, headers, output_rows)
    return path


def export_pre_transforms(excel_path: Path, output_dir: Path) -> Path:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, PRE_TRANSFORMS_SHEET, required=True)
    assert sheet is not None

    rows = sheet_rows(sheet)

    # Допускаем пустой лист с двумя строками шапки и без данных
    if len(rows) < 2:
        raise ValueError("Pre-transforms sheet must contain at least double header")

    root = output_dir / PRE_TRANSFORMS_DIR
    ensure_dir(root)

    # если данных нет, просто возвращаем пустой каталог
    if len(rows) == 2:
        return root

    for row in rows[2:]:
        values = [normalize_cell(v) for v in row[:5]]

        while len(values) < 5:
            values.append("")

        if is_row_empty(values):
            continue

        target_table, source_tables_raw, transformation_sql, comments, settings_sql = values

        if not target_table:
            raise ValueError("Pre-transforms row has empty target table")

        target_dir = root / slugify_dir_name(target_table)
        ensure_dir(target_dir)

        payload = {
            "target_table": target_table,
            "source_tables": split_lines(source_tables_raw),
            "comments": comments or None,
        }

        write_json_file(target_dir / "pre-transform.json", payload)

        if transformation_sql:
            write_text_file(target_dir / "preliminary_transformation.sql", transformation_sql)

        if settings_sql:
            write_text_file(target_dir / "settings.sql", settings_sql)

    return root


def export_joins(excel_path: Path, output_dir: Path) -> Path:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, JOINS_SHEET, required=True)
    assert sheet is not None
    rows = sheet_rows(sheet)
    if len(rows) < 3:
        raise ValueError("Joins must contain double header and data rows")

    root = output_dir / JOINS_DIR
    ensure_dir(root)

    for row in rows[2:]:
        values = [normalize_cell(v) for v in row[:10]]
        while len(values) < 10:
            values.append("")
        if is_row_empty(values):
            continue

        (
            load_code,
            table_name,
            description,
            table_codes_raw,
            delta_codes_raw,
            source_tables_join_sql,
            load_code_params_raw,
            settings_table_join_sql,
            history_rule,
            business_history_dates,
        ) = values

        if not load_code or not table_name:
            raise ValueError("Joins row must contain table_name and load_code")

        join_dir = root / slugify_dir_name(table_name) / slugify_dir_name(load_code)
        ensure_dir(join_dir)

        payload = {
            "description": description or None,
            "table_codes": split_lines(table_codes_raw),
            "table_codes_to_track_delta": split_lines(delta_codes_raw),
            "load_code_params": split_lines(load_code_params_raw),
            "history_rule": history_rule or None,
            "business_history_dates": business_history_dates or None,
        }
        payload = {k: v for k, v in payload.items() if v not in (None, [], "")}
        write_json_file(join_dir / "join.json", payload)

        if source_tables_join_sql:
            write_text_file(join_dir / "source_tables_join.sql", source_tables_join_sql)
        if settings_table_join_sql:
            write_text_file(join_dir / "settings_table_join.sql", settings_table_join_sql)

    return root


def export_mappings(excel_path: Path, output_dir: Path) -> tuple[Path, Path]:
    ensure_dir(output_dir)

    sheet = get_sheet(excel_path, MAPPINGS_SHEET, required=True)
    assert sheet is not None
    rows = sheet_rows(sheet)
    if len(rows) < 2:
        raise ValueError("Mappings sheet must contain header and data rows")

    joins_root = output_dir / JOINS_DIR
    ensure_dir(joins_root)

    grouped_rows: dict[tuple[str, str], list[dict[str, str]]] = {}
    attribute_names_by_table: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))

    for row in rows[2:]:
        values = [normalize_cell(v) for v in row[:7]]
        while len(values) < 7:
            values.append("")
        if is_row_empty(values):
            continue

        load_code, table_name, attribute_code, attribute_name, mapping_algorithm, additional_join, settings = values

        if not load_code or not table_name or not attribute_code:
            raise ValueError("Mappings row must contain load_code, table_name, attribute_code")

        grouped_rows.setdefault((table_name, load_code), []).append(
            {
                "attribute_code": attribute_code,
                "mapping_algorithm": mapping_algorithm,
                "additional_join": additional_join,
                "settings": settings,
            }
        )

        if attribute_name:
            attribute_names_by_table[table_name][attribute_code].append(attribute_name)

    for (table_name, load_code), mapping_rows in grouped_rows.items():
        mapping_dir = joins_root / slugify_dir_name(table_name) / slugify_dir_name(load_code)
        ensure_dir(mapping_dir)

        csv_rows: list[list[str]] = []
        extra_payload: dict[str, dict[str, str]] = {}

        for item in mapping_rows:
            csv_rows.append([item["attribute_code"], item["mapping_algorithm"]])

            extra: dict[str, str] = {}
            if item["additional_join"]:
                extra["additional_join"] = item["additional_join"]
            if item["settings"]:
                extra["settings"] = item["settings"]

            if extra:
                extra_payload[item["attribute_code"]] = extra

        write_csv_rows(mapping_dir / MAPPINGS_CSV, ["attribute_code", "mapping_algorithm"], csv_rows)

        if extra_payload:
            write_json_file(mapping_dir / MAPPINGS_EXTRA_JSON, extra_payload)

    resolved_table_names: dict[str, dict[str, str]] = {}
    for table_name, attrs in attribute_names_by_table.items():
        resolved_table_names[table_name] = {}
        for attribute_code, names in attrs.items():
            unique_names = {n for n in names if n}
            if len(unique_names) > 1:
                raise ValueError(
                    f"Conflicting attribute_name inside table '{table_name}' "
                    f"for attribute_code '{attribute_code}': {sorted(unique_names)}"
                )
            if unique_names:
                resolved_table_names[table_name][attribute_code] = next(iter(unique_names))

    attribute_to_table_values: dict[str, dict[str, str]] = defaultdict(dict)
    for table_name, attrs in resolved_table_names.items():
        for attribute_code, attribute_name in attrs.items():
            attribute_to_table_values[attribute_code][table_name] = attribute_name

    common_names: dict[str, str] = {}
    table_overrides: dict[str, dict[str, str]] = {}

    for attribute_code, table_values in attribute_to_table_values.items():
        unique_values = set(table_values.values())
        if len(unique_values) == 1:
            common_names[attribute_code] = next(iter(unique_values))
        else:
            for table_name, attribute_name in table_values.items():
                table_overrides.setdefault(table_name, {})[attribute_code] = attribute_name

    attribute_names_payload = {
        "common": dict(sorted(common_names.items())),
        "tables": {
            table_name: dict(sorted(attrs.items()))
            for table_name, attrs in sorted(table_overrides.items())
        },
    }

    attribute_names_path = output_dir / ATTRIBUTE_NAMES_JSON
    write_json_file(attribute_names_path, attribute_names_payload)

    return joins_root, attribute_names_path


def section_key_from_title(title: str) -> str:
    return (
        title.strip()
        .lower()
        .replace("-", " ")
        .replace("/", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace(":", "")
        .replace(".", "")
        .replace(",", "")
        .replace("  ", " ")
        .replace(" ", "_")
    )

def export_excel_to_repo(excel_path: str, output_dir: str) -> None:
    excel_file = Path(excel_path)
    out_dir = Path(output_dir)

    created: list[Path] = []
    created.append(export_change_history(excel_file, out_dir))
    created.append(export_source_lg(excel_file, out_dir))
    created.append(export_targets(excel_file, out_dir))
    created.append(export_pre_transforms(excel_file, out_dir))
    created.append(export_joins(excel_file, out_dir))

    mappings_root, attribute_names_path = export_mappings(excel_file, out_dir)
    created.append(mappings_root)
    created.append(attribute_names_path)

    for maybe_path in [
        export_simple_csv_sheet(excel_file, out_dir, SETTINGS_SHEET, SETTINGS_CSV, SETTINGS_HEADERS, required=False),
        export_simple_csv_sheet(excel_file, out_dir, PARAMETERS_SHEET, PARAMETERS_CSV, PARAMETERS_HEADERS, required=False),
        export_simple_csv_sheet(excel_file, out_dir, ST_DECODER_SHEET, ST_DECODER_CSV, ST_DECODER_HEADERS, required=False),
        export_simple_csv_sheet(excel_file, out_dir, ST_FILTER_SHEET, ST_FILTER_CSV, ST_FILTER_HEADERS, required=False),
    ]:
        if maybe_path is not None:
            created.append(maybe_path)

    for path in created:
        print(f"Created: {path}")


if __name__ == "__main__":
    export_excel_to_repo("input.xlsx", "./repo")