from __future__ import annotations

import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

from s2t_tool.adapters.excel.writer import build_rich_diff, maybe_build_rich_diff
from s2t_tool.adapters.excel.writer_style import append_csv_sheet, finalize_sheet_style
from s2t_tool.shared.csv_files import write_csv_rows


class ExcelDiffOutputTests(unittest.TestCase):
    def test_returns_rich_text_for_changed_value(self) -> None:
        result = build_rich_diff("before", "after")
        self.assertIsInstance(result, CellRichText)
        self.assertTrue(any(isinstance(part, TextBlock) for part in result))

    def test_returns_plain_text_when_values_are_equal(self) -> None:
        self.assertEqual(build_rich_diff("same", "same"), "same")

    def test_non_diff_mode_returns_new_text(self) -> None:
        self.assertEqual(maybe_build_rich_diff(False, "before", "after"), "after")

    def test_sql_like_change_produces_rich_text(self) -> None:
        old = "when x in ('a','b') then 1\nwhen x = 'c' then 2"
        new = "when x = 'a' then 1\nwhen x in ('b','c') then 2"
        result = build_rich_diff(old, new)
        self.assertIsInstance(result, CellRichText)

    def test_append_csv_sheet_highlights_changed_data_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            current = Path(temp_dir) / "current.csv"
            old = Path(temp_dir) / "old.csv"
            write_csv_rows(current, ["value"], [["after"]])
            write_csv_rows(old, ["value"], [["before"]])

            wb = Workbook()
            wb.remove(wb.active)
            append_csv_sheet(
                wb=wb,
                title="Settings",
                csv_path=current,
                config={"global": {}},
                required=True,
                pre_transforms_sheet="Pre-transforms",
                joins_sheet="Joins",
                mappings_sheet="Mappings",
                diff_csv_path=old,
                maybe_build_rich_diff=maybe_build_rich_diff,
            )

            sheet = wb["Settings"]
            self.assertIsInstance(sheet["A2"].value, CellRichText)

    def test_finalize_sheet_style_adds_fill_for_diff_cells(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Joins"
        ws["A1"] = "Header"
        ws["A2"] = build_rich_diff("before", "after")

        finalize_sheet_style(
            ws,
            config={"global": {}},
            sheet_name="Joins",
            pre_transforms_sheet="Pre-transforms",
            joins_sheet="Joins",
            mappings_sheet="Mappings",
        )

        self.assertEqual(ws["A2"].fill.fill_type, "solid")

    def test_saved_rich_text_uses_opaque_argb_colors(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "diff.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = build_rich_diff("before", "after")
            wb.save(output)

            with zipfile.ZipFile(output) as archive:
                xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")

            self.assertIn('rgb="FFFF0000"', xml)
            self.assertIn('rgb="FF008000"', xml)


if __name__ == "__main__":
    unittest.main()
