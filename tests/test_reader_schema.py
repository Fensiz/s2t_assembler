from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from s2t_tool.domain.schema import DEFAULT_SCHEMA
from s2t_tool.infrastructure.excel_reader import ExcelRepoReader


class ReaderSchemaTests(unittest.TestCase):
    def test_reader_resolves_sheet_aliases(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            excel_path = root / "input.xlsx"
            output_dir = root / "out"

            wb = Workbook()
            sheet = wb.active
            sheet.title = "ChangeHistory"
            sheet.append(["Author", "Date", "Version", "Description", "Jira ticket"])
            wb.save(excel_path)

            reader = ExcelRepoReader(excel_path=excel_path, output_dir=output_dir)
            resolved = reader.get_sheet(DEFAULT_SCHEMA.change_history_sheet)

            self.assertIsNotNone(resolved)
            assert resolved is not None
            self.assertEqual(resolved.title, "ChangeHistory")

    def test_reader_exports_optional_csv_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            excel_path = root / "input.xlsx"
            output_dir = root / "out"

            wb = Workbook()
            wb.remove(wb.active)
            sheet = wb.create_sheet("Settings")
            sheet.append(["Settings alias", "Settings description", "Settings table", "Settings type", "Period", "Mask"])
            sheet.append(["a", "b", "c", "d", "e", "f"])
            wb.save(excel_path)

            reader = ExcelRepoReader(excel_path=excel_path, output_dir=output_dir)
            path = reader.export_simple_csv_sheet(
                sheet_name=DEFAULT_SCHEMA.settings_sheet,
                output_name=DEFAULT_SCHEMA.settings_csv,
                headers=list(DEFAULT_SCHEMA.settings_headers),
                required=False,
            )

            self.assertIsNotNone(path)
            assert path is not None
            self.assertTrue(path.exists())
            self.assertIn("a;b;c;d;e;f", path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
