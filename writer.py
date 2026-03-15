from __future__ import annotations

import json
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from common import *


CHANGE_HISTORY_JSON = "change-history.json"
SOURCE_LG_CSV = "source-lg.csv"
TARGETS_CSV = "targets.csv"
SETTINGS_CSV = "settings.csv"
PARAMETERS_CSV = "parameters.csv"
ST_DECODER_CSV = "st_decoder.csv"
ST_FILTER_CSV = "st_filter.csv"
METADATA_JSON = "metadata.json"
ATTRIBUTE_NAMES_JSON = "attribute_names.json"

PRE_TRANSFORMS_DIR = "pre-transforms"
JOINS_DIR = "joins"

CHANGE_HISTORY_SHEET = "Change history"
SOURCE_LG_SHEET = "Source LG"
TARGETS_SHEET = "Targets"
PRE_TRANSFORMS_SHEET = "Pre-transforms"
JOINS_SHEET = "Joins"
MAPPINGS_SHEET = "Mappings"
SETTINGS_SHEET = "Settings"
PARAMETERS_SHEET = "Parameters"
ST_DECODER_SHEET = "ST_DECODER"
ST_FILTER_SHEET = "ST_FILTER"
METADATA_SHEET = "Metadata"

WRITER_CONFIG_FILE = "writer_config.json"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

EXCEL_HEADER_LABELS = {
    "scheme": "Scheme",
    "table": "Table",
    "column": "Column",
    "data_type": "Data type",
    "data_length": "Data length",
    "is_key": "Is a key",
    "description": "Description",
    "link": "Link",
    "table_code": "Table code",
    "table_name": "Table name",
    "settings_alias": "Settings alias",
    "settings_description": "Settings description",
    "settings_table": "Settings table",
    "settings_type": "Settings type",
    "period": "Period",
    "mask": "Mask",
    "parameter": "Parameter",
    "value": "Value",
    "comment": "Comment",
    "source_value": "Source value",
    "target_value": "Target value",
    "start_dt": "Start dt",
    "end_dt": "End dt",
    "update_date": "Update date",
    "author_update_name": "Author update name",
    "author": "Author",
    "date": "Date",
    "version": "Version",
    "jira_ticket": "Jira ticket",
    "load_code": "Load code",
    "attribute_code": "Attribute code",
    "attribute_name": "Attribute name",
    "mapping_algorithm": "Mapping algorithm",
    "additional_join": "Additional join",
    "business_history_dates": "Business history dates",
    "load_code_params": "Load code params",
    "table_codes_to_track_delta": "Table codes to track delta",
    "source_tables_join": "Source tables join",
    "settings_table_join": "Settings table join",
    "history_rule": "History rule",
}

BLACK_FONT = InlineFont(color="000000")
GREEN_FONT = InlineFont(color="008000")
RED_FONT = InlineFont(color="FF0000", strike=True)


# ============================================================
# Diff helpers
# ============================================================

def build_rich_diff(old_text: str | None, new_text: str | None) -> CellRichText | str:
    old_value = "" if old_text is None else str(old_text)
    new_value = "" if new_text is None else str(new_text)

    if old_value == new_value:
        return new_value

    matcher = SequenceMatcher(a=old_value, b=new_value)
    rich = CellRichText()

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        old_chunk = old_value[i1:i2]
        new_chunk = new_value[j1:j2]

        if tag == "equal":
            if old_chunk:
                rich.append(old_chunk)
        elif tag == "delete":
            if old_chunk:
                rich.append(TextBlock(RED_FONT, old_chunk))
        elif tag == "insert":
            if new_chunk:
                rich.append(TextBlock(GREEN_FONT, new_chunk))
        elif tag == "replace":
            if old_chunk:
                rich.append(TextBlock(RED_FONT, old_chunk))
            if new_chunk:
                rich.append(TextBlock(GREEN_FONT, new_chunk))

    return rich


def maybe_build_rich_diff(
    diff_enabled: bool,
    old_value: str | None,
    new_value: str | None,
) -> str | CellRichText:
    """
    Return plain text in normal mode and rich diff in diff mode.
    """
    new_text = "" if new_value is None else str(new_value)

    if not diff_enabled:
        return new_text

    return build_rich_diff(old_value, new_text)


def normalize_key_part(value: object) -> str:
    return "" if value is None else str(value)


def join_row_key(table_name: str, load_code: str) -> tuple[str, str]:
    return (
        normalize_key_part(table_name),
        normalize_key_part(load_code),
    )


def pre_transform_row_key(target_table: str) -> str:
    return normalize_key_part(target_table)


def mapping_row_key(load_code: str, table_name: str, attribute_code: str) -> tuple[str, str, str]:
    return (
        normalize_key_part(load_code),
        normalize_key_part(table_name),
        normalize_key_part(attribute_code),
    )


# ============================================================
# Generic helpers
# ============================================================

def repo_header_to_excel(header: str) -> str:
    return EXCEL_HEADER_LABELS.get(header, header.replace("_", " ").capitalize())


def load_writer_config(config_path: str | Path | None = None) -> dict[str, Any]:
    if config_path is None:
        return load_json_resource(WRITER_CONFIG_FILE)

    path = Path(config_path)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))

    return load_json_resource(str(config_path))


def create_sheet(wb: Workbook, title: str) -> Worksheet:
    return wb.create_sheet(title=title)


def sheet_config(config: dict[str, Any], sheet_name: str) -> dict[str, Any]:
    global_cfg = config.get("global", {})
    per_sheet = config.get("sheets", {}).get(sheet_name, {})
    return {**global_cfg, **per_sheet}


def color_fill(hex_color: str | None) -> PatternFill | None:
    if not hex_color:
        return None
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)


def center_merged_headers(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    for merge_range in cfg.get("header_merge_ranges", []):
        first_cell = merge_range.split(":")[0]
        cell = sheet[first_cell]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_base_style(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    font_name = cfg.get("font_name", "Calibri")
    font_size = cfg.get("font_size", 11)
    wrap_text = cfg.get("wrap_text", True)
    vertical_alignment = cfg.get("vertical_alignment", "top")

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(name=font_name, size=font_size)
            cell.alignment = Alignment(wrap_text=wrap_text, vertical=vertical_alignment)


def apply_header_style(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    header_rows = cfg.get("header_rows", 1)
    fill = color_fill(cfg.get("header_fill_color"))
    font_name = cfg.get("font_name", "Calibri")
    font_size = cfg.get("font_size", 11)

    for row_idx in range(1, header_rows + 1):
        for cell in sheet[row_idx]:
            cell.font = Font(name=font_name, size=font_size, bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill


def estimate_row_height_for_text(
    values: list[str],
    min_height: float,
    max_height: float,
    line_height: float = 15.0,
) -> float:
    max_lines = 1
    for value in values:
        if not value:
            continue
        max_lines = max(max_lines, str(value).count("\n") + 1)

    estimated = max(min_height, max_lines * line_height)
    return min(estimated, max_height)


def row_has_multiline_values(sheet: Worksheet, row_idx: int) -> bool:
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=row_idx, column=col_idx).value
        if value is not None and "\n" in str(value):
            return True
    return False


def row_values_as_strings(sheet: Worksheet, row_idx: int) -> list[str]:
    values: list[str] = []
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=row_idx, column=col_idx).value
        values.append("" if value is None else str(value))
    return values


def apply_row_heights(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    header_rows = cfg.get("header_rows", 1)
    header_height = cfg.get("header_row_height", 24)

    row_height_mode = cfg.get("row_height_mode", "fixed")
    body_row_height = cfg.get("body_row_height", 36)
    auto_height_on_multiline = cfg.get("auto_height_on_multiline", False)
    min_body_row_height = cfg.get("min_body_row_height", body_row_height)
    max_body_row_height = cfg.get("max_body_row_height", 180)

    for row_idx in range(1, sheet.max_row + 1):
        if row_idx <= header_rows:
            sheet.row_dimensions[row_idx].height = header_height
            continue

        if row_height_mode == "auto":
            values = row_values_as_strings(sheet, row_idx)
            sheet.row_dimensions[row_idx].height = estimate_row_height_for_text(
                values=values,
                min_height=min_body_row_height,
                max_height=max_body_row_height,
            )
            continue

        # fixed mode
        if auto_height_on_multiline and row_has_multiline_values(sheet, row_idx):
            values = row_values_as_strings(sheet, row_idx)
            sheet.row_dimensions[row_idx].height = estimate_row_height_for_text(
                values=values,
                min_height=min_body_row_height,
                max_height=max_body_row_height,
            )
        else:
            sheet.row_dimensions[row_idx].height = body_row_height


def apply_column_widths(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    widths = cfg.get("column_widths", {})
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width


def apply_merge_ranges(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    for merge_range in cfg.get("header_merge_ranges", []):
        sheet.merge_cells(merge_range)


def apply_freeze_panes(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    freeze_panes = cfg.get("freeze_panes")
    if freeze_panes:
        sheet.freeze_panes = freeze_panes


def apply_autofilter(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    if not cfg.get("add_autofilter", True):
        return

    if sheet.max_column == 0 or sheet.max_row == 0:
        return

    autofilter_row = cfg.get("autofilter_row", cfg.get("header_rows", 1))
    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A{autofilter_row}:{last_col}{sheet.max_row}"


def fill_row(sheet: Worksheet, row_idx: int, color_hex: str | None) -> None:
    fill = color_fill(color_hex)
    if fill is None:
        return

    for col_idx in range(1, sheet.max_column + 1):
        sheet.cell(row=row_idx, column=col_idx).fill = fill


def apply_special_header_colors(sheet: Worksheet, sheet_name: str, cfg: dict[str, Any]) -> None:
    group_colors = cfg.get("group_colors", {})
    if not group_colors:
        return

    loadcode_fill = color_fill(group_colors.get("loadcode_header"))
    target_fill = color_fill(group_colors.get("target_section"))
    source_fill = color_fill(group_colors.get("source_section"))

    if sheet_name == PRE_TRANSFORMS_SHEET:
        for col in range(1, 3):
            sheet.cell(row=1, column=col).fill = target_fill
        for col in range(3, 6):
            sheet.cell(row=1, column=col).fill = source_fill

        for col in range(1, 3):
            sheet.cell(row=2, column=col).fill = target_fill
        for col in range(3, 6):
            sheet.cell(row=2, column=col).fill = source_fill

    elif sheet_name == JOINS_SHEET:
        sheet.cell(row=1, column=1).fill = loadcode_fill
        for col in range(2, 6):
            sheet.cell(row=1, column=col).fill = target_fill
        for col in range(6, 11):
            sheet.cell(row=1, column=col).fill = source_fill

        sheet.cell(row=2, column=1).fill = loadcode_fill
        for col in range(2, 6):
            sheet.cell(row=2, column=col).fill = target_fill
        for col in range(6, 11):
            sheet.cell(row=2, column=col).fill = source_fill

    elif sheet_name == MAPPINGS_SHEET:
        sheet.cell(row=1, column=1).fill = loadcode_fill
        for col in range(2, 5):
            sheet.cell(row=1, column=col).fill = target_fill
        for col in range(5, 8):
            sheet.cell(row=1, column=col).fill = source_fill

        sheet.cell(row=2, column=1).fill = loadcode_fill
        for col in range(2, 5):
            sheet.cell(row=2, column=col).fill = target_fill
        for col in range(5, 8):
            sheet.cell(row=2, column=col).fill = source_fill


def apply_cell_borders(sheet: Worksheet) -> None:
    for row in sheet.iter_rows(
        min_row=1,
        max_row=sheet.max_row,
        min_col=1,
        max_col=sheet.max_column,
    ):
        for cell in row:
            cell.border = THIN_BORDER


def apply_table_borders(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: int = 2,
) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row=row_idx, column=col_idx).border = THIN_BORDER


def apply_alternating_group_fill(sheet: Worksheet, cfg: dict[str, Any]) -> None:
    grouping = cfg.get("grouping", {})
    if not grouping.get("enabled", False):
        return

    group_by_column = grouping.get("group_by_column")
    if not group_by_column:
        return

    group_col_idx = column_index_from_string(group_by_column)
    header_rows = cfg.get("header_rows", 1)
    alt_color = cfg.get("group_colors", {}).get("alternating_fill")
    if not alt_color:
        return

    previous_group_value = None
    group_index = 0

    for row_idx in range(header_rows + 1, sheet.max_row + 1):
        current_value = sheet.cell(row=row_idx, column=group_col_idx).value

        if current_value != previous_group_value:
            group_index += 1
            previous_group_value = current_value

        if group_index % 2 == 0:
            fill_row(sheet, row_idx, alt_color)


def finalize_sheet_style(sheet: Worksheet, config: dict[str, Any], sheet_name: str) -> None:
    cfg = sheet_config(config, sheet_name)
    apply_base_style(sheet, cfg)
    apply_header_style(sheet, cfg)
    apply_merge_ranges(sheet, cfg)
    center_merged_headers(sheet, cfg)
    apply_special_header_colors(sheet, sheet_name, cfg)
    apply_row_heights(sheet, cfg)
    apply_column_widths(sheet, cfg)
    apply_freeze_panes(sheet, cfg)
    apply_autofilter(sheet, cfg)
    apply_alternating_group_fill(sheet, cfg)
    if cfg.get("apply_global_borders", True):
        apply_cell_borders(sheet)


# ============================================================
# Generic sheet builders
# ============================================================

def append_csv_sheet(
    wb: Workbook,
    title: str,
    csv_path: Path,
    config: dict[str, Any],
    required: bool = False,
) -> None:
    if not csv_path.exists():
        if required:
            raise ValueError(f"Required CSV not found: {csv_path}")
        return

    sheet = create_sheet(wb, title)
    headers, rows = read_csv_rows(csv_path)

    if headers:
        pretty_headers = [repo_header_to_excel(h) for h in headers]
        sheet.append(pretty_headers)

    for row in rows:
        sheet.append(row)

    finalize_sheet_style(sheet, config, title)


def build_change_history_sheet(wb: Workbook, repo_dir: Path, config: dict[str, Any]) -> None:
    path = repo_dir / CHANGE_HISTORY_JSON
    entries = read_json_file(path, default=[])
    sheet = create_sheet(wb, CHANGE_HISTORY_SHEET)

    sheet.append(["Author", "Date", "Version", "Description", "Jira ticket"])

    for entry in entries:
        sheet.append(
            [
                entry.get("author"),
                entry.get("date"),
                entry.get("version"),
                entry.get("description"),
                entry.get("jira_ticket"),
            ]
        )

    finalize_sheet_style(sheet, config, CHANGE_HISTORY_SHEET)


def build_source_lg_sheet(wb: Workbook, repo_dir: Path, config: dict[str, Any]) -> None:
    append_csv_sheet(wb, SOURCE_LG_SHEET, repo_dir / SOURCE_LG_CSV, config, required=True)


def build_targets_sheet(wb: Workbook, repo_dir: Path, config: dict[str, Any]) -> None:
    append_csv_sheet(wb, TARGETS_SHEET, repo_dir / TARGETS_CSV, config, required=True)


# ============================================================
# Pre-transforms
# ============================================================

def build_pre_transforms_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    diff_enabled = diff_repo_dir is not None

    sheet = create_sheet(wb, PRE_TRANSFORMS_SHEET)

    sheet.cell(row=1, column=1, value="TARGET")
    sheet.cell(row=1, column=3, value="SOURCE")

    headers = [
        "Target table",
        "Source tables",
        "Transformation SQL",
        "Comment",
        "Settings SQL",
    ]

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    root = repo_dir / PRE_TRANSFORMS_DIR

    if not root.exists():
        finalize_sheet_style(sheet, config, PRE_TRANSFORMS_SHEET)
        return

    old_rows_by_key: dict[str, dict[str, str]] = {}

    if diff_enabled:
        old_root = Path(diff_repo_dir) / PRE_TRANSFORMS_DIR  # type: ignore[arg-type]
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                json_path = old_table_dir / "pre-transform.json"
                if not json_path.exists():
                    continue

                payload = read_json_file(json_path, default={}) or {}
                target_table = payload.get("target_table", old_table_dir.name)

                old_rows_by_key[pre_transform_row_key(target_table)] = {
                    "source_tables": "\n".join(payload.get("source_tables", [])),
                    "comments": payload.get("comments", ""),
                    "transformation_sql": read_text_file(old_table_dir / "preliminary_transformation.sql"),
                    "settings_sql": read_text_file(old_table_dir / "settings.sql"),
                }

    row_idx = 3

    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        json_path = table_dir / "pre-transform.json"
        if not json_path.exists():
            continue

        payload = read_json_file(json_path, default={}) or {}

        target_table = payload.get("target_table", table_dir.name)
        source_tables = "\n".join(payload.get("source_tables", []))
        comments = payload.get("comments", "")

        transformation_sql = read_text_file(table_dir / "preliminary_transformation.sql")
        settings_sql = read_text_file(table_dir / "settings.sql")

        key = pre_transform_row_key(target_table)
        old_row = old_rows_by_key.get(key, {})

        sheet.cell(row=row_idx, column=1, value=target_table)
        sheet.cell(
            row=row_idx,
            column=2,
            value=maybe_build_rich_diff(diff_enabled, old_row.get("source_tables"), source_tables),
        )
        sheet.cell(
            row=row_idx,
            column=3,
            value=maybe_build_rich_diff(diff_enabled, old_row.get("transformation_sql"), transformation_sql),
        )
        sheet.cell(
            row=row_idx,
            column=4,
            value=maybe_build_rich_diff(diff_enabled, old_row.get("comments"), comments),
        )
        sheet.cell(
            row=row_idx,
            column=5,
            value=maybe_build_rich_diff(diff_enabled, old_row.get("settings_sql"), settings_sql),
        )

        row_idx += 1

    finalize_sheet_style(sheet, config, PRE_TRANSFORMS_SHEET)


# ============================================================
# Joins
# ============================================================

def build_joins_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    diff_enabled = diff_repo_dir is not None

    root = repo_dir / JOINS_DIR
    if not root.exists():
        raise ValueError(f"Required directory not found: {root}")

    old_rows_by_key: dict[tuple[str, str], dict[str, str]] = {}

    if diff_enabled:
        old_root = Path(diff_repo_dir) / JOINS_DIR  # type: ignore[arg-type]
        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                for old_load_dir in sorted([p for p in old_table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
                    join_json = read_json_file(old_load_dir / "join.json", default={}) or {}

                    key = join_row_key(old_table_dir.name, old_load_dir.name)
                    old_rows_by_key[key] = {
                        "description": join_json.get("description", ""),
                        "table_codes": "\n".join(join_json.get("table_codes", [])),
                        "table_codes_to_track_delta": "\n".join(join_json.get("table_codes_to_track_delta", [])),
                        "source_tables_join_sql": read_text_file(old_load_dir / "source_tables_join.sql"),
                        "load_code_params": "\n".join(join_json.get("load_code_params", [])),
                        "settings_table_join_sql": read_text_file(old_load_dir / "settings_table_join.sql"),
                        "history_rule": join_json.get("history_rule", ""),
                        "business_history_dates": join_json.get("business_history_dates", ""),
                    }

    sheet = create_sheet(wb, JOINS_SHEET)
    sheet["B1"] = "TARGET"
    sheet["F1"] = "SOURCE"

    headers = [
        "Load code",
        "Table name",
        "Description",
        "Table code(s)",
        "Table codes to track delta",
        "Source tables join",
        "Load code params",
        "Settings table join",
        "History rule",
        "Business history dates",
    ]

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    row_idx = 3

    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        for load_dir in sorted([p for p in table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
            join_json = read_json_file(load_dir / "join.json", default={}) or {}

            key = join_row_key(table_dir.name, load_dir.name)
            old_row = old_rows_by_key.get(key, {})

            current_description = join_json.get("description", "")
            current_table_codes = "\n".join(join_json.get("table_codes", []))
            current_delta_codes = "\n".join(join_json.get("table_codes_to_track_delta", []))
            current_source_join_sql = read_text_file(load_dir / JOIN_SOURCE_SQL)
            current_load_code_params = "\n".join(join_json.get("load_code_params", []))
            current_settings_join_sql = read_text_file(load_dir / "settings_table_join.sql")
            current_history_rule = join_json.get("history_rule", "")
            current_business_history_dates = join_json.get("business_history_dates", "")

            sheet.cell(row=row_idx, column=1, value=load_dir.name)
            sheet.cell(row=row_idx, column=2, value=table_dir.name)
            sheet.cell(
                row=row_idx,
                column=3,
                value=maybe_build_rich_diff(diff_enabled, old_row.get("description"), current_description),
            )
            sheet.cell(
                row=row_idx,
                column=4,
                value=maybe_build_rich_diff(diff_enabled, old_row.get("table_codes"), current_table_codes),
            )
            sheet.cell(
                row=row_idx,
                column=5,
                value=maybe_build_rich_diff(
                    diff_enabled,
                    old_row.get("table_codes_to_track_delta"),
                    current_delta_codes,
                ),
            )
            sheet.cell(
                row=row_idx,
                column=6,
                value=maybe_build_rich_diff(
                    diff_enabled,
                    old_row.get("source_tables_join_sql"),
                    current_source_join_sql,
                ),
            )
            sheet.cell(
                row=row_idx,
                column=7,
                value=maybe_build_rich_diff(diff_enabled, old_row.get("load_code_params"), current_load_code_params),
            )
            sheet.cell(
                row=row_idx,
                column=8,
                value=maybe_build_rich_diff(
                    diff_enabled,
                    old_row.get("settings_table_join_sql"),
                    current_settings_join_sql,
                ),
            )
            sheet.cell(
                row=row_idx,
                column=9,
                value=maybe_build_rich_diff(diff_enabled, old_row.get("history_rule"), current_history_rule),
            )
            sheet.cell(
                row=row_idx,
                column=10,
                value=maybe_build_rich_diff(
                    diff_enabled,
                    old_row.get("business_history_dates"),
                    current_business_history_dates,
                ),
            )

            row_idx += 1

    finalize_sheet_style(sheet, config, JOINS_SHEET)


# ============================================================
# Mappings
# ============================================================

def resolve_attribute_name(table_name: str, attribute_code: str, attribute_names: dict[str, Any]) -> str:
    tables = attribute_names.get("tables", {}) or {}
    common = attribute_names.get("common", {}) or {}

    if table_name in tables and attribute_code in tables[table_name]:
        return tables[table_name][attribute_code]
    return common.get(attribute_code, "")


def build_mappings_sheet(
    wb: Workbook,
    repo_dir: Path,
    config: dict[str, Any],
    diff_repo_dir: str | None = None,
) -> None:
    diff_enabled = diff_repo_dir is not None

    old_rows_by_key: dict[tuple[str, str, str], dict[str, str]] = {}

    if diff_enabled:
        old_root = Path(diff_repo_dir) / JOINS_DIR  # type: ignore[arg-type]
        old_attribute_names = read_json_file(Path(diff_repo_dir) / ATTRIBUTE_NAMES_JSON, default={}) or {}

        if old_root.exists():
            for old_table_dir in sorted([p for p in old_root.iterdir() if p.is_dir()], key=lambda p: p.name):
                for old_load_dir in sorted([p for p in old_table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
                    mappings_path = old_load_dir / "mappings.csv"
                    if not mappings_path.exists():
                        continue

                    _, old_rows = read_csv_rows(mappings_path)
                    old_extra = read_json_file(old_load_dir / "mappings.extra.json", default={}) or {}

                    for row in old_rows:
                        while len(row) < 2:
                            row.append("")

                        attribute_code = row[0]
                        mapping_algorithm = row[1]
                        extra_item = old_extra.get(attribute_code, {}) or {}

                        key = mapping_row_key(old_load_dir.name, old_table_dir.name, attribute_code)
                        old_rows_by_key[key] = {
                            "attribute_name": resolve_attribute_name(
                                old_table_dir.name,
                                attribute_code,
                                old_attribute_names,
                            ),
                            "mapping_algorithm": mapping_algorithm,
                            "additional_join": extra_item.get("additional_join", ""),
                            "settings": extra_item.get("settings", ""),
                        }

    root = repo_dir / JOINS_DIR
    attribute_names = read_json_file(repo_dir / ATTRIBUTE_NAMES_JSON, default={}) or {}

    sheet = create_sheet(wb, MAPPINGS_SHEET)

    # 1 строка — секции
    sheet.cell(row=1, column=2, value="TARGET")
    sheet.cell(row=1, column=5, value="SOURCE")

    # 2 строка — заголовки колонок
    headers = [
        "Load code",
        "Table name",
        "Attribute code",
        "Attribute name",
        "Mapping algorithm",
        "Additional join",
        "Settings",
    ]

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

    # данные начинаются строго с 3 строки
    row_idx = 3

    for table_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name):
        for load_dir in sorted([p for p in table_dir.iterdir() if p.is_dir()], key=lambda p: p.name):
            mappings_path = load_dir / "mappings.csv"
            if not mappings_path.exists():
                continue

            _, rows = read_csv_rows(mappings_path)
            extra = read_json_file(load_dir / "mappings.extra.json", default={}) or {}

            for row in rows:
                while len(row) < 2:
                    row.append("")

                attribute_code = row[0]
                mapping_algorithm = normalize_newlines(row[1])
                extra_item = extra.get(attribute_code, {}) or {}

                sheet.cell(row=row_idx, column=1, value=load_dir.name)
                sheet.cell(row=row_idx, column=2, value=table_dir.name)
                sheet.cell(row=row_idx, column=3, value=attribute_code)

                key = mapping_row_key(load_dir.name, table_dir.name, attribute_code)
                old_row = old_rows_by_key.get(key, {})

                current_attribute_name = resolve_attribute_name(table_dir.name, attribute_code, attribute_names)
                current_additional_join = extra_item.get("additional_join", "")
                current_settings = extra_item.get("settings", "")

                sheet.cell(
                    row=row_idx,
                    column=4,
                    value=maybe_build_rich_diff(diff_enabled, old_row.get("attribute_name"), current_attribute_name),
                )
                sheet.cell(
                    row=row_idx,
                    column=5,
                    value=maybe_build_rich_diff(diff_enabled, old_row.get("mapping_algorithm"), mapping_algorithm),
                )
                sheet.cell(
                    row=row_idx,
                    column=6,
                    value=maybe_build_rich_diff(diff_enabled, old_row.get("additional_join"), current_additional_join),
                )
                sheet.cell(
                    row=row_idx,
                    column=7,
                    value=maybe_build_rich_diff(diff_enabled, old_row.get("settings"), current_settings),
                )

                row_idx += 1

    finalize_sheet_style(sheet, config, MAPPINGS_SHEET)


# ============================================================
# Metadata
# ============================================================

def normalize_table_row(values: Any, headers: list[str] | None) -> list[Any]:
    if isinstance(values, list):
        return values

    if isinstance(values, dict):
        if headers:
            return [values.get(header, "") for header in headers]
        return list(values.values())

    return [values]


def build_metadata_sheet(
    wb: Workbook,
    config: dict[str, Any],
    diff_commit: str | None = None,
) -> None:
    payload = load_json_resource(METADATA_JSON) or {}
    blocks = payload.get("blocks", [])

    sheet = create_sheet(wb, METADATA_SHEET)
    row_idx = 1

    if diff_commit:
        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        cell = sheet.cell(
            row=row_idx,
            column=1,
            value=f"DIFF FILE — DO NOT USE FOR PUT (compared with commit: {diff_commit})",
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_idx += 2

    for block in blocks:
        title_en = block.get("title_en")
        title_ru = block.get("title_ru")
        link = block.get("link")
        table = block.get("table")

        # English title
        if title_en:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=title_en)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

        # Russian title
        if title_ru:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=title_ru)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row_idx += 1

        # Optional plain text / link
        if link:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=link)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        # Optional extra text lines
        for text_line in block.get("lines", []) or []:
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=text_line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1

        # Optional table
        if table:
            headers = table.get("headers")
            rows = table.get("rows", []) or []

            table_start_row = row_idx

            header_names: list[str] | None = None
            if headers:
                header_names = [str(h) for h in headers]

                for col_idx, header in enumerate(header_names, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                row_idx += 1

            max_width = len(header_names) if header_names else 0

            for raw_row in rows:
                normalized_row = normalize_table_row(raw_row, header_names)
                max_width = max(max_width, len(normalized_row))

                for col_idx, value in enumerate(normalized_row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

                row_idx += 1

            table_end_row = row_idx - 1

            if max_width > 0 and table_end_row >= table_start_row:
                apply_table_borders(sheet, table_start_row, table_end_row, 1, max_width)

        row_idx += 1

    finalize_sheet_style(sheet, config, METADATA_SHEET)


# ============================================================
# Main entry
# ============================================================

def build_excel_from_repo(
    repo_dir: str,
    output_excel_path: str,
    config_path: str = WRITER_CONFIG_FILE,
    diff_repo_dir: str | None = None,
    diff_commit: str | None = None,
) -> None:
    repo = Path(repo_dir)
    config = load_writer_config(config_path)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    build_change_history_sheet(wb, repo, config)
    build_source_lg_sheet(wb, repo, config)
    build_targets_sheet(wb, repo, config)
    build_pre_transforms_sheet(wb, repo, config, diff_repo_dir)
    build_joins_sheet(wb, repo, config, diff_repo_dir)
    build_mappings_sheet(wb, repo, config, diff_repo_dir)

    append_csv_sheet(wb, SETTINGS_SHEET, repo / SETTINGS_CSV, config, required=False)
    append_csv_sheet(wb, PARAMETERS_SHEET, repo / PARAMETERS_CSV, config, required=False)
    append_csv_sheet(wb, ST_DECODER_SHEET, repo / ST_DECODER_CSV, config, required=False)
    append_csv_sheet(wb, ST_FILTER_SHEET, repo / ST_FILTER_CSV, config, required=False)

    build_metadata_sheet(wb, config, diff_commit)

    wb.save(output_excel_path)
    print(f"Created: {output_excel_path}")


if __name__ == "__main__":
    build_excel_from_repo("./repo", "output.xlsx", WRITER_CONFIG_FILE)